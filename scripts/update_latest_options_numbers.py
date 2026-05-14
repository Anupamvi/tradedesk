#!/usr/bin/env python3
"""Build a Schwab-backed options P&L sheet in a Numbers workbook.

The workbook is a free-form Numbers ledger, so this script creates or replaces
one clean reconciliation sheet instead of appending duplicate loose rows.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sqlite3
import subprocess
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_START_DATE = date(2026, 4, 1)
DEFAULT_SHEET_NAME = "Codex Schwab Numbers"
LEGACY_GENERATED_SHEETS = ["Schwab Apr-May 2026"]


@dataclass(frozen=True)
class Execution:
    ticker: str
    symbol: str
    expiry: str
    position_effect: str
    quantity: float
    price: float
    net_amount: float
    executed_at: str
    order_id: str

    @property
    def dt(self) -> datetime:
        return parse_dt(self.executed_at)

    @property
    def is_opening(self) -> bool:
        return self.position_effect.upper() == "OPENING"

    @property
    def side(self) -> str:
        if self.is_opening:
            return "Buy" if self.net_amount < 0 else "Sell"
        return "Sell to close" if self.net_amount > 0 else "Buy to close"


def parse_dt(value: str) -> datetime:
    cleaned = value.replace("+0000", "+00:00")
    return datetime.fromisoformat(cleaned)


def parse_date(value: str) -> date | None:
    if not value:
        return None
    return parse_dt(value).date()


def money(value: float | None) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    sign = "-" if value < 0 else ""
    return f"{sign}${abs(value):,.0f}"


def cashflow(value: float) -> str:
    if abs(value) < 0.005:
        return "$0"
    label = "Credit" if value > 0 else "Debit"
    return f"{label} {money(abs(value))}"


def normalize_key_text(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", value.upper()).strip()


def option_parts(symbol: str) -> tuple[str, str, float]:
    # OCC-ish Schwab symbol: "NVDA  260508C00185000"
    compact = re.sub(r"\s+", "", symbol)
    m = re.match(r"([A-Z.]+)(\d{6})([CP])(\d{8})$", compact)
    if not m:
        return "", "", 0.0
    yymmdd = m.group(2)
    expiry = f"20{yymmdd[:2]}-{yymmdd[2:4]}-{yymmdd[4:6]}"
    strike = int(m.group(4)) / 1000
    return expiry, m.group(3), strike


def leg_label(ex: Execution) -> str:
    expiry, cp, strike = option_parts(ex.symbol)
    strike_txt = f"{strike:g}"
    qty_txt = f"{abs(ex.quantity):g}"
    return f"{ex.side} {qty_txt} {ex.ticker} {expiry} ${strike_txt}{cp}"


def compact_leg_label(ex: Execution) -> str:
    expiry, cp, strike = option_parts(ex.symbol)
    qty_txt = f"{abs(ex.quantity):g}"
    return f"{ex.side} {qty_txt} {ex.ticker} {expiry} ${strike:g}{cp}"


def strategy_for(open_legs: list[Execution]) -> str:
    if len(open_legs) == 1:
        expiry, cp, _strike = option_parts(open_legs[0].symbol)
        side = open_legs[0].side
        if side == "Sell" and cp == "P":
            return "Short Put"
        if side == "Sell" and cp == "C":
            return "Short Call"
        if side == "Buy" and cp == "P":
            return "Long Put"
        if side == "Buy" and cp == "C":
            return "Long Call"
        return "Single Option"

    parsed = []
    for ex in open_legs:
        expiry, cp, strike = option_parts(ex.symbol)
        parsed.append((ex, expiry, cp, strike, ex.side))
    expiries = {p[1] for p in parsed}
    cps = {p[2] for p in parsed}
    if len(open_legs) == 2 and len(expiries) == 1 and len(cps) == 1:
        cp = parsed[0][2]
        buy = [p for p in parsed if p[4] == "Buy"]
        sell = [p for p in parsed if p[4] == "Sell"]
        if len(buy) == 1 and len(sell) == 1:
            buy_strike = buy[0][3]
            sell_strike = sell[0][3]
            if cp == "C":
                return "Bull Call Debit" if buy_strike < sell_strike else "Bear Call Credit"
            if cp == "P":
                return "Bear Put Debit" if buy_strike > sell_strike else "Bull Put Credit"
    if len(open_legs) == 4 and cps == {"C", "P"}:
        return "Defined-Risk Condor/Combo"
    return "Multi-Leg Option"


def load_executions(db_path: Path) -> list[Execution]:
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True, timeout=10)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        """
        select ticker, symbol, expiry, position_effect, quantity, price,
               net_amount, executed_at, order_id
        from raw_executions
        where symbol is not null
          and symbol <> ''
          and expiry is not null
          and expiry <> ''
          and executed_at >= '2026-03-01'
        order by executed_at, order_id, symbol
        """
    ).fetchall()
    con.close()
    seen: set[tuple[Any, ...]] = set()
    out: list[Execution] = []
    for row in rows:
        key = tuple(row[k] for k in row.keys())
        if key in seen:
            continue
        seen.add(key)
        out.append(Execution(**dict(row)))
    return out


def load_closed_trades(schwab_json: Path) -> dict[str, dict[str, Any]]:
    data = json.loads(schwab_json.read_text())
    closed_by_entry: dict[str, dict[str, Any]] = {}
    for trade in data["accounts"][0].get("closed_trades", []):
        for order_id in trade.get("entry_order_ids", []):
            closed_by_entry[str(order_id)] = trade
    return closed_by_entry


def load_positions(schwab_json: Path) -> dict[str, dict[str, Any]]:
    data = json.loads(schwab_json.read_text())
    out = {}
    for pos in data["accounts"][0].get("positions", []):
        if pos.get("asset_type") == "OPTION":
            out[pos["symbol"]] = pos
    return out


def load_manual_text(export_xlsx: Path) -> str:
    wb = load_workbook(export_xlsx, data_only=True)
    texts: list[str] = []
    if "Options" not in wb.sheetnames:
        return ""
    ws = wb["Options"]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                texts.append(cell.value)
    return normalize_key_text(" ".join(texts))


def manual_match(legs: list[Execution], manual_text: str) -> str:
    if not manual_text:
        return "not checked"
    hits = 0
    for leg in legs:
        expiry, cp, strike = option_parts(leg.symbol)
        exp_dt = datetime.strptime(expiry, "%Y-%m-%d").date() if expiry else None
        expiry_alts = [expiry]
        if exp_dt:
            expiry_alts.extend(
                [
                    exp_dt.strftime("%m/%d/%Y"),
                    exp_dt.strftime("%m/%d/%y"),
                    exp_dt.strftime("%m %d %Y"),
                    exp_dt.strftime("%m %d %y"),
                ]
            )
        strike_alts = [f"{strike:g}", f"{strike:.2f}", f"{int(strike)} 00" if strike.is_integer() else ""]
        needles = []
        for exp in expiry_alts:
            for strike_txt in strike_alts:
                if exp and strike_txt:
                    needles.append(normalize_key_text(f"{leg.ticker} {exp} {strike_txt} {cp}"))
        needles.extend(
            normalize_key_text(x)
            for x in [
                f"{leg.ticker} {strike:g}{cp}",
                f"{leg.ticker} {strike:.2f} {cp}",
                f"{leg.ticker} {int(strike)} 00 {cp}" if strike.is_integer() else "",
            ]
        )
        if any(needle and needle in manual_text for needle in needles):
            hits += 1
    if hits == len(legs):
        return "already in manual log"
    if hits:
        return "partial manual match"
    return "not found in manual log"


def build_rows(
    executions: list[Execution],
    closed_by_entry: dict[str, dict[str, Any]],
    positions: dict[str, dict[str, Any]],
    manual_text: str,
    start_date: date,
    end_date: date,
) -> list[dict[str, Any]]:
    openings_by_order: dict[str, list[Execution]] = defaultdict(list)
    closes_by_symbol: dict[str, list[Execution]] = defaultdict(list)
    for ex in executions:
        if ex.is_opening:
            d = ex.dt.date()
            if start_date <= d <= end_date:
                openings_by_order[ex.order_id].append(ex)
        else:
            closes_by_symbol[ex.symbol].append(ex)

    rows: list[dict[str, Any]] = []
    today = datetime.now(timezone.utc).date()

    for order_id, legs in sorted(openings_by_order.items(), key=lambda item: min(x.dt for x in item[1])):
        legs = sorted(legs, key=lambda x: (x.ticker, x.expiry, x.symbol, x.net_amount))
        opened_dt = min(x.dt for x in legs)
        ticker = legs[0].ticker if len({x.ticker for x in legs}) == 1 else "MULTI"
        expiries = sorted({option_parts(x.symbol)[0] or x.expiry for x in legs})
        expiry_txt = ", ".join(expiries)
        entry_cash = sum(x.net_amount for x in legs)
        closed = closed_by_entry.get(order_id)
        open_pnl = 0.0
        realized_pnl: float | None = None
        notes: list[str] = []
        open_leg_labels: list[str] = []
        closed_leg_labels: list[str] = []

        if closed:
            status = "CLOSED PROFIT" if closed.get("realized_pnl", 0) >= 0 else "CLOSED LOSS"
            realized_pnl = float(closed.get("realized_pnl", 0) or 0)
            closed_at = closed.get("closed_at", "")
            closed_leg_labels = [compact_leg_label(x) for x in legs]
        else:
            closed_at = ""
            realized_partial = 0.0
            partial_notes: list[str] = []
            has_current = False
            all_expired_absent = True
            for leg in legs:
                pos = positions.get(leg.symbol)
                if pos:
                    has_current = True
                    all_expired_absent = False
                    open_pnl += float(pos.get("total_pnl") or 0)
                    open_leg_labels.append(compact_leg_label(leg))
                expiry_date = parse_date((option_parts(leg.symbol)[0] or leg.expiry) + "T00:00:00+00:00")
                if pos or not expiry_date or expiry_date >= today:
                    all_expired_absent = False
                leg_close_cash = 0.0
                leg_closes: list[Execution] = []
                for close in closes_by_symbol.get(leg.symbol, []):
                    if close.dt > opened_dt:
                        leg_close_cash += close.net_amount
                        leg_closes.append(close)
                if not pos and abs(leg_close_cash) > 0.005:
                    realized_partial += leg.net_amount + leg_close_cash
                    closed_leg_labels.append(
                        f"{compact_leg_label(leg)} -> "
                        + " + ".join(compact_leg_label(x) for x in leg_closes)
                    )
                    partial_notes.append(
                        f"{leg.ticker} {option_parts(leg.symbol)[2]:g}{option_parts(leg.symbol)[1]} "
                        f"entry {money(leg.net_amount)}, close {money(leg_close_cash)}"
                    )
                elif not pos and expiry_date and expiry_date < today:
                    realized_partial += leg.net_amount
                    closed_leg_labels.append(f"{compact_leg_label(leg)} -> expired/absent")
            if has_current:
                status = "OPEN"
                if abs(realized_partial) > 0.005:
                    status = "PARTIAL OPEN"
                    realized_pnl = realized_partial
                    notes.append(
                        "partial: closed-leg P/L = "
                        + money(realized_partial)
                        + "; open-leg mark = "
                        + money(open_pnl)
                    )
                    if partial_notes:
                        notes.append("; ".join(partial_notes))
            elif all_expired_absent:
                status = "CLOSED PROFIT" if entry_cash >= 0 else "CLOSED LOSS"
                realized_pnl = entry_cash
                closed_at = f"expired {expiry_txt}"
                closed_leg_labels = [compact_leg_label(x) for x in legs]
            else:
                status = "UNKNOWN"
                realized_pnl = entry_cash + realized_partial if realized_partial else None
                notes.append("no current position and no reconciled close found")

        total_pnl = (realized_pnl or 0.0) + open_pnl
        if status == "OPEN":
            total_pnl = open_pnl
        if status == "PARTIAL OPEN":
            pass

        rows.append(
            {
                "opened": opened_dt.date().isoformat(),
                "status": status,
                "ticker": ticker,
                "strategy": strategy_for(legs),
                "expiry": expiry_txt,
                "legs": " / ".join(leg_label(x) for x in legs),
                "open_legs": " / ".join(open_leg_labels),
                "closed_legs": " / ".join(closed_leg_labels),
                "entry_cash": cashflow(entry_cash),
                "realized_pnl": money(realized_pnl) if realized_pnl is not None else "",
                "open_pnl": money(open_pnl) if abs(open_pnl) > 0.005 else "",
                "total_pnl": money(total_pnl),
                "closed_or_asof": closed_at or datetime.now().date().isoformat(),
                "order_ids": order_id,
                "manual_match": manual_match(legs, manual_text),
                "notes": "; ".join(notes),
                "sort_pnl": total_pnl,
            }
        )
    return rows


def applescript_string(value: Any) -> str:
    text = "" if value is None else str(value)
    return '"' + text.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ") + '"'


def month_label(opened: str) -> str:
    # Avoid Numbers auto-coercing bare "April 2026" into April 1, 2026.
    return datetime.strptime(opened, "%Y-%m-%d").strftime("%B %Y Trades")


def status_mix(rows: list[dict[str, Any]]) -> str:
    statuses = sorted({row["status"] for row in rows})
    return ", ".join(f"{status}={sum(1 for row in rows if row['status'] == status)}" for status in statuses)


def pnl_total(rows: list[dict[str, Any]]) -> float:
    return sum(float(row["sort_pnl"]) for row in rows if row.get("sort_pnl") is not None)


def write_numbers_sheet(workbook: Path, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    headers = [
        "Opened",
        "Status",
        "Ticker",
        "Strategy",
        "Expiry",
        "Legs",
        "Open Legs",
        "Closed Legs",
        "Entry Cashflow",
        "Realized P/L",
        "Open P/L",
        "Total P/L",
        "Closed / As Of",
        "Order IDs",
        "Manual Log Match",
        "Notes",
    ]
    table: list[list[Any]] = []
    section_rows: list[int] = []
    header_rows: list[int] = []
    status_rows: list[tuple[int, str]] = []
    key_map = {
        "Opened": "opened",
        "Status": "status",
        "Ticker": "ticker",
        "Strategy": "strategy",
        "Expiry": "expiry",
        "Legs": "legs",
        "Open Legs": "open_legs",
        "Closed Legs": "closed_legs",
        "Entry Cashflow": "entry_cash",
        "Realized P/L": "realized_pnl",
        "Open P/L": "open_pnl",
        "Total P/L": "total_pnl",
        "Closed / As Of": "closed_or_asof",
        "Order IDs": "order_ids",
        "Manual Log Match": "manual_match",
        "Notes": "notes",
    }
    current_month = ""
    for row in rows:
        label = month_label(row["opened"])
        if label != current_month:
            current_month = label
            section_rows.append(len(table) + 1)
            table.append([label] + [""] * (len(headers) - 1))
            header_rows.append(len(table) + 1)
            table.append(headers)
        status_rows.append((len(table) + 1, row["status"]))
        table.append([row[key_map[h]] for h in headers])
    if not table:
        table.append(["No Schwab option trades found"] + [""] * (len(headers) - 1))

    script_lines = [
        f"set workbookPath to {applescript_string(str(workbook))}",
        f"set targetSheetName to {applescript_string(sheet_name)}",
        "tell application \"Numbers\"",
        "\topen (POSIX file workbookPath)",
        "\tdelay 1",
        "\ttell front document",
        "\t\ttry",
        "\t\t\tdelete sheet targetSheetName",
        "\t\ton error",
        "\t\tend try",
        f"\t\tset legacyGeneratedSheets to {{{', '.join(applescript_string(x) for x in LEGACY_GENERATED_SHEETS)}}}",
        "\t\trepeat with legacyName in legacyGeneratedSheets",
        "\t\t\tif (legacyName as text) is not targetSheetName then",
        "\t\t\t\ttry",
        "\t\t\t\t\tdelete sheet (legacyName as text)",
        "\t\t\t\ton error",
        "\t\t\t\tend try",
        "\t\t\tend if",
        "\t\tend repeat",
        f"\t\tset newSheet to make new sheet at end of sheets with properties {{name:{applescript_string(sheet_name)}}}",
        "\t\ttell newSheet",
        "\t\t\ttell table 1",
        f"\t\t\t\tset row count to {len(table)}",
        f"\t\t\t\tset column count to {len(headers)}",
        f"\t\t\t\tset name to {applescript_string(sheet_name)}",
        "\t\t\t\tset header row count to 0",
    ]
    for r, row in enumerate(table, start=1):
        for c, value in enumerate(row, start=1):
            script_lines.append(f"\t\t\t\tset value of cell {c} of row {r} to {applescript_string(value)}")
    for r in section_rows:
        script_lines.extend(
            [
                f"\t\t\t\tset background color of row {r} to {{15750, 15750, 15750}}",
                f"\t\t\t\tset text color of row {r} to {{65535, 65535, 65535}}",
            ]
        )
    for r in header_rows:
        script_lines.extend(
            [
                f"\t\t\t\tset background color of row {r} to {{52428, 52428, 52428}}",
                f"\t\t\t\tset text color of row {r} to {{65535, 65535, 65535}}",
            ]
        )
    for r, status in status_rows:
        if status == "OPEN":
            color = "{65535, 62258, 36044}"
        elif status == "PARTIAL OPEN":
            color = "{65535, 55769, 34438}"
        elif "PROFIT" in status:
            color = "{48742, 64764, 48742}"
        elif "LOSS" in status:
            color = "{65535, 49344, 49344}"
        else:
            color = "{61166, 61166, 61166}"
        script_lines.append(f"\t\t\t\tset background color of row {r} to {color}")
    script_lines.extend(
        [
            "\t\t\tend tell",
            "\t\tend tell",
            "\t\tsave",
            "\tend tell",
            "end tell",
        ]
    )
    subprocess.run(["osascript", "-e", "\n".join(script_lines)], check=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--export-xlsx", required=True, type=Path)
    parser.add_argument("--schwab-json", required=True, type=Path)
    parser.add_argument("--state-db", required=True, type=Path)
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--start-date", default=DEFAULT_START_DATE.isoformat())
    parser.add_argument("--end-date", default=datetime.now().date().isoformat())
    parser.add_argument("--out-json", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    start_date = date.fromisoformat(args.start_date)
    end_date = date.fromisoformat(args.end_date)
    executions = load_executions(args.state_db)
    closed_by_entry = load_closed_trades(args.schwab_json)
    positions = load_positions(args.schwab_json)
    manual = load_manual_text(args.export_xlsx)
    rows = build_rows(executions, closed_by_entry, positions, manual, start_date, end_date)
    rows.sort(key=lambda r: (r["opened"], r["ticker"], r["order_ids"]))

    if args.out_json:
        args.out_json.write_text(json.dumps(rows, indent=2))

    print(
        f"Built {len(rows)} option trade rows from Schwab API/state "
        f"for {start_date.isoformat()} -> {end_date.isoformat()}."
    )
    current_month = end_date.strftime("%Y-%m")
    current_month_rows = [row for row in rows if str(row.get("opened", "")).startswith(current_month)]

    print("Full-range status mix:", status_mix(rows))
    print("Full-range Total P/L:", money(pnl_total(rows)))
    print("Current month:", current_month)
    print("Current-month rows:", len(current_month_rows))
    print("Current-month status mix:", status_mix(current_month_rows) if current_month_rows else "none")
    print("Current-month Total P/L:", money(pnl_total(current_month_rows)))
    if not args.dry_run:
        write_numbers_sheet(args.workbook, args.sheet_name, rows)
        print(f"Updated Numbers sheet: {args.sheet_name}")


if __name__ == "__main__":
    main()
