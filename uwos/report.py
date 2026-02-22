#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
from collections import defaultdict, deque
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import requests
import yaml

DEFAULT_SHEET_CACHE_CSV = r"c:\uw_root\out\playbook\source_sheet_latest.csv"
DEFAULT_SHEET_REALIZED_CSV = r"c:\uw_root\out\playbook\cleaned_realized_trades_from_sheet.csv"
SCHWAB_OPTION_SYMBOL_RE = re.compile(r"^\s*([A-Z\. ]{1,6})(\d{6})([CP])(\d{8})\s*$")
COMPACT_OPTION_SYMBOL_RE = re.compile(r"^\s*([A-Z\.]{1,6})(\d{6})([CP])(\d{8})\s*$")


def norm_col(name: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "_", str(name).strip().lower())
    return re.sub(r"_+", "_", s).strip("_")


def parse_float(x) -> float:
    if x is None:
        return math.nan
    if isinstance(x, (int, float, np.number)):
        try:
            return float(x)
        except Exception:
            return math.nan
    s = str(x).strip()
    if not s:
        return math.nan
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").replace("%", "")
    if s.startswith("-"):
        neg = True
    try:
        v = float(s)
        return -abs(v) if neg else v
    except Exception:
        return math.nan


def find_col(columns: List[str], aliases: List[str]) -> Optional[str]:
    mapped = {c: norm_col(c) for c in columns}
    for alias in aliases:
        a = norm_col(alias)
        for raw, n in mapped.items():
            if n == a:
                return raw
    return None


def compute_profit_factor(pnl: pd.Series) -> float:
    gp = float(pnl[pnl > 0].sum())
    gl = float(-pnl[pnl < 0].sum())
    if gl <= 0:
        return math.inf if gp > 0 else math.nan
    return gp / gl


def compute_max_drawdown(pnl: pd.Series) -> float:
    if pnl.empty:
        return 0.0
    c = pnl.cumsum()
    dd = c - c.cummax()
    return float(dd.min())


def longest_streak(flags: List[bool], target: bool) -> int:
    best = 0
    cur = 0
    for f in flags:
        if bool(f) == target:
            cur += 1
            if cur > best:
                best = cur
        else:
            cur = 0
    return best


def trailing_loss_streak(pnl: pd.Series) -> int:
    k = 0
    for v in reversed(pnl.tolist()):
        if v < 0:
            k += 1
        else:
            break
    return k


def load_config(path: Path) -> Dict:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}
    return cfg if isinstance(cfg, dict) else {}


def load_realized_trades(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, low_memory=False)
    cols = list(df.columns)
    date_col = find_col(cols, ["date", "trade_date", "close_date", "closed_at"])
    pnl_col = find_col(cols, ["realized_pnl", "pnl", "profit_loss", "net_pnl", "amount"])
    strategy_col = find_col(cols, ["strategy", "strategy_type", "setup"])
    symbol_col = find_col(cols, ["symbol", "ticker", "underlying", "underlying_symbol"])

    if not date_col or not pnl_col:
        raise ValueError(f"Missing required columns in {path}. Need date + realized pnl.")

    out = pd.DataFrame()
    out["date"] = pd.to_datetime(df[date_col], errors="coerce")
    out["realized_pnl"] = df[pnl_col].map(parse_float)
    out["strategy"] = df[strategy_col].astype(str).str.strip() if strategy_col else "Unknown"
    out["symbol"] = df[symbol_col].astype(str).str.strip().str.upper() if symbol_col else "UNKNOWN"
    out["symbol"] = out["symbol"].replace({"": "UNKNOWN", "NAN": "UNKNOWN"})
    out["strategy"] = out["strategy"].replace({"": "Unknown", "nan": "Unknown"})

    out = out.dropna(subset=["date", "realized_pnl"]).copy()
    out = out.sort_values("date").reset_index(drop=True)
    out["month"] = out["date"].dt.to_period("M").astype(str)
    out["weekday"] = out["date"].dt.day_name()
    out["win"] = out["realized_pnl"] > 0
    return out


def parse_sheet_id_and_gid(sheet_csv_url: str) -> Tuple[str, str]:
    url = str(sheet_csv_url or "").strip()
    m_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9\-_]+)", url)
    if not m_id:
        raise ValueError(f"Could not parse Google Sheet ID from URL: {sheet_csv_url}")
    m_gid = re.search(r"[?&]gid=(\d+)", url)
    return m_id.group(1), (m_gid.group(1) if m_gid else "")


def normalize_headers(values: List[Any]) -> List[str]:
    headers: List[str] = []
    seen: Dict[str, int] = {}
    for i, v in enumerate(values):
        raw = str(v).strip() if v is not None else ""
        h = raw if raw else f"Unnamed: {i}"
        if h in seen:
            seen[h] += 1
            h = f"{h}.{seen[h]}"
        else:
            seen[h] = 0
        headers.append(h)
    return headers


def _hex_to_rgb(hex_text: str) -> Optional[Tuple[int, int, int]]:
    s = str(hex_text or "").strip().upper()
    if not s:
        return None
    s = s.replace("#", "")
    if len(s) == 8:
        s = s[2:]
    if not re.fullmatch(r"[0-9A-F]{6}", s):
        return None
    return int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)


def _is_yellow_rgb(rgb: Tuple[int, int, int]) -> bool:
    r, g, b = rgb
    return r >= 225 and g >= 185 and b <= 150 and (r - b) >= 90 and (g - b) >= 50


def _cell_is_yellow(cell: Any) -> bool:
    fill = getattr(cell, "fill", None)
    if fill is None or not getattr(fill, "fill_type", None):
        return False
    colors = [getattr(fill, "fgColor", None), getattr(fill, "start_color", None)]
    for color in colors:
        rgb = _hex_to_rgb(getattr(color, "rgb", ""))
        if rgb and _is_yellow_rgb(rgb):
            return True
    return False


def _is_manual_section_header(text: Any) -> bool:
    s = "" if text is None else str(text).strip()
    if not s:
        return False
    try:
        from uwos.analyze_trading_year import parse_manual_section_date  # local module

        return pd.notna(parse_manual_section_date(s))
    except Exception:
        # Fallback: Month-YY / Month YYYY
        return bool(
            re.match(
                r"^\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s*[- ]\s*(\d{2,4})\s*$",
                s,
                flags=re.I,
            )
        )


def load_sheet_rows_from_xlsx(sheet_csv_url: str, row_filter: str) -> Tuple[pd.DataFrame, Dict[str, float]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError(
            "Row-color filtering requires openpyxl. Install with: python -m pip install openpyxl"
        ) from e

    sheet_id, gid = parse_sheet_id_and_gid(sheet_csv_url)
    xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    if gid:
        xlsx_url += f"&gid={gid}&single=true"

    resp = requests.get(xlsx_url, timeout=60)
    resp.raise_for_status()
    wb = load_workbook(BytesIO(resp.content), data_only=True)
    ws = wb.active

    rows: List[List[Any]] = []
    yellow_flags: List[bool] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        vals = [cell.value for cell in row]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        rows.append(vals)
        yellow_flags.append(any(_cell_is_yellow(cell) for cell in row))

    if not rows:
        raise RuntimeError("XLSX export loaded but contained no non-empty rows.")

    headers = normalize_headers(rows[0])
    data_rows = rows[1:]
    data_yellow = yellow_flags[1:]

    kept_rows: List[List[Any]] = []
    yellow_kept = 0
    for vals, is_yellow in zip(data_rows, data_yellow):
        keep = True
        if row_filter == "yellow":
            c0 = vals[0] if vals else None
            keep = bool(is_yellow) or _is_manual_section_header(c0)
        if keep:
            kept_rows.append(vals)
            if is_yellow:
                yellow_kept += 1

    out = pd.DataFrame(kept_rows, columns=headers)
    meta = {
        "xlsx_nonempty_rows": float(len(rows)),
        "xlsx_data_rows": float(len(data_rows)),
        "xlsx_total_yellow_rows": float(sum(1 for f in data_yellow if f)),
        "xlsx_kept_rows": float(len(kept_rows)),
        "xlsx_kept_yellow_rows": float(yellow_kept),
    }
    return out, meta


def build_realized_from_sheet_csv_url(
    sheet_csv_url: str,
    raw_cache_csv: Path,
    realized_out_csv: Path,
    row_filter: str = "all",
) -> Dict[str, float]:
    """
    Pull Google Sheet CSV export URL, parse manual options log rows into standardized realized trades,
    and persist both raw cache + standardized realized CSV.
    """
    raw_cache_csv.parent.mkdir(parents=True, exist_ok=True)
    realized_out_csv.parent.mkdir(parents=True, exist_ok=True)

    ingest_meta: Dict[str, float] = {}
    if str(row_filter).strip().lower() == "yellow":
        raw_df, ingest_meta = load_sheet_rows_from_xlsx(sheet_csv_url, row_filter="yellow")
    else:
        raw_df = pd.read_csv(sheet_csv_url, low_memory=False)
        ingest_meta = {"csv_rows_loaded": float(len(raw_df))}
    raw_df.to_csv(raw_cache_csv, index=False)

    try:
        from uwos.analyze_trading_year import standardize_manual_options_log_df  # local module
    except Exception as e:
        raise RuntimeError(f"Failed to import manual log parser from analyze_trading_year.py: {e}") from e

    std_df, meta = standardize_manual_options_log_df(raw_df, source_name=sheet_csv_url)
    if std_df.empty:
        raise RuntimeError("Sheet CSV parsed but no usable realized rows were extracted.")
    std_df.to_csv(realized_out_csv, index=False)
    out_meta = dict(meta or {})
    out_meta.update(ingest_meta)
    return out_meta


def option_strategy_label(symbol: str, open_side: str) -> str:
    _, right, _ = parse_option_meta(symbol)
    if open_side == "SHORT":
        return "Short Put Option" if right == "P" else "Short Call Option"
    return "Long Put Option" if right == "P" else "Long Call Option"


def build_realized_from_schwab_api(realized_out_csv: Path, lookback_days: int = 365) -> Dict[str, float]:
    try:
        from uwos.schwab_auth import SchwabAuthConfig, SchwabLiveDataService
    except Exception as e:
        raise RuntimeError(f"Failed to import Schwab live service: {e}") from e

    realized_out_csv.parent.mkdir(parents=True, exist_ok=True)

    cfg_live = SchwabAuthConfig.from_env(load_dotenv_file=True)
    svc = SchwabLiveDataService(cfg_live)
    svc.connect()
    cli = svc._client

    acct_resp = cli.get_account_numbers()
    acct_resp.raise_for_status()
    acct_rows = acct_resp.json()
    account_hashes = [a.get("hashValue") for a in acct_rows if a.get("hashValue")]
    if not account_hashes:
        raise RuntimeError("No Schwab account hashes available.")

    end_dt = dt.datetime.now(dt.timezone.utc)
    start_dt = end_dt - dt.timedelta(days=max(1, int(lookback_days)))

    raw_events: List[Dict[str, Any]] = []
    for acc_hash in account_hashes:
        resp = cli.get_transactions(
            acc_hash,
            start_date=start_dt,
            end_date=end_dt,
            transaction_types=[cli.Transactions.TransactionType.TRADE],
        )
        resp.raise_for_status()
        for txn in resp.json():
            txn_time = pd.to_datetime(txn.get("time"), errors="coerce")
            if pd.isna(txn_time):
                continue
            for it in txn.get("transferItems", []) or []:
                inst = it.get("instrument", {}) or {}
                if str(inst.get("assetType", "")).upper() != "OPTION":
                    continue
                symbol = str(inst.get("symbol", "")).strip().upper()
                if not symbol:
                    continue
                pos_eff = str(it.get("positionEffect", "")).strip().upper()
                if pos_eff not in {"OPENING", "CLOSING"}:
                    continue
                qty_signed = parse_float(it.get("amount"))
                qty_abs = abs(qty_signed) if np.isfinite(qty_signed) else math.nan
                cost = parse_float(it.get("cost"))
                if (not np.isfinite(qty_abs)) or qty_abs <= 0 or (not np.isfinite(cost)):
                    continue
                raw_events.append(
                    {
                        "account_hash": acc_hash,
                        "time": txn_time,
                        "symbol": symbol,
                        "position_effect": pos_eff,
                        "qty_signed": float(qty_signed),
                        "qty_abs": float(qty_abs),
                        "cost": float(cost),
                    }
                )

    if not raw_events:
        raise RuntimeError("No option trade events returned from Schwab transactions.")

    events = sorted(raw_events, key=lambda x: x["time"])
    # FIFO lots by account+symbol and side opened.
    lots = defaultdict(lambda: {"LONG": deque(), "SHORT": deque()})
    closed_rows: List[Dict[str, Any]] = []

    for ev in events:
        acc = ev["account_hash"]
        sym = ev["symbol"]
        qty = float(ev["qty_abs"])
        qty_signed = float(ev["qty_signed"])
        cost = float(ev["cost"])
        t = pd.to_datetime(ev["time"])
        eff = ev["position_effect"]
        unit_cash = cost / max(1e-9, qty)

        if eff == "OPENING":
            open_side = "LONG" if qty_signed > 0 else "SHORT"
            lots[(acc, sym)][open_side].append(
                {
                    "qty": qty,
                    "unit_cash": unit_cash,
                    "open_date": t,
                    "open_side": open_side,
                }
            )
            continue

        # CLOSING: buy-to-close -> qty_signed>0 closes SHORT; sell-to-close -> qty_signed<0 closes LONG.
        close_side = "SHORT" if qty_signed > 0 else "LONG"
        qleft = qty
        qbook = lots[(acc, sym)][close_side]
        while qleft > 1e-9 and qbook:
            lot = qbook[0]
            m = min(qleft, float(lot["qty"]))
            realized = (float(lot["unit_cash"]) + unit_cash) * m
            underlying = option_underlying_symbol(sym) or sym
            strategy = option_strategy_label(sym, str(lot["open_side"]))
            closed_rows.append(
                {
                    "source_file": "schwab_api",
                    "date": t,
                    "open_date": pd.to_datetime(lot["open_date"]),
                    "symbol": underlying,
                    "strategy": strategy,
                    "side": str(lot["open_side"]),
                    "qty": m,
                    "entry_price": math.nan,
                    "exit_price": math.nan,
                    "fees": 0.0,
                    "dte": math.nan,
                    "conviction": math.nan,
                    "realized_pnl": realized,
                }
            )
            lot["qty"] = float(lot["qty"]) - m
            qleft -= m
            if lot["qty"] <= 1e-9:
                qbook.popleft()

    out = pd.DataFrame(closed_rows)
    if out.empty:
        raise RuntimeError("No closed option lots could be reconstructed from Schwab transaction events.")
    out = out.sort_values("date").reset_index(drop=True)
    out["holding_days"] = (pd.to_datetime(out["date"]) - pd.to_datetime(out["open_date"])).dt.days
    out["win"] = out["realized_pnl"] > 0
    keep = [
        "source_file",
        "date",
        "open_date",
        "symbol",
        "strategy",
        "side",
        "qty",
        "entry_price",
        "exit_price",
        "fees",
        "dte",
        "conviction",
        "realized_pnl",
        "holding_days",
        "win",
    ]
    out[keep].to_csv(realized_out_csv, index=False)
    return {
        "schwab_accounts": float(len(account_hashes)),
        "schwab_option_events": float(len(events)),
        "schwab_closed_lots": float(len(out)),
    }


def build_open_option_lot_metadata_from_schwab_api(
    cli: Any,
    account_number_to_hash: Dict[str, str],
    lookback_days: int = 365,
) -> pd.DataFrame:
    end_dt = dt.datetime.now(dt.timezone.utc)
    start_dt = end_dt - dt.timedelta(days=max(1, int(lookback_days)))
    raw_events: List[Dict[str, Any]] = []
    for acct_num, acc_hash in account_number_to_hash.items():
        if not acc_hash:
            continue
        resp = cli.get_transactions(
            acc_hash,
            start_date=start_dt,
            end_date=end_dt,
            transaction_types=[cli.Transactions.TransactionType.TRADE],
        )
        resp.raise_for_status()
        for txn in resp.json():
            t = pd.to_datetime(txn.get("time"), errors="coerce")
            if pd.isna(t):
                continue
            for it in txn.get("transferItems", []) or []:
                inst = it.get("instrument", {}) or {}
                if str(inst.get("assetType", "")).upper() != "OPTION":
                    continue
                sym = str(inst.get("symbol", "")).strip().upper()
                if not sym:
                    continue
                pos_eff = str(it.get("positionEffect", "")).strip().upper()
                if pos_eff not in {"OPENING", "CLOSING"}:
                    continue
                amt = parse_float(it.get("amount"))
                qty_abs = abs(amt) if np.isfinite(amt) else math.nan
                if not np.isfinite(qty_abs) or qty_abs <= 0:
                    continue
                raw_events.append(
                    {
                        "account_number": str(acct_num).strip(),
                        "symbol": sym,
                        "time": t,
                        "position_effect": pos_eff,
                        "qty_signed": float(amt),
                        "qty_abs": float(qty_abs),
                    }
                )
    if not raw_events:
        return pd.DataFrame(columns=["account_number", "symbol", "side", "open_qty_est", "open_date_est", "last_open_trade_time"])

    events = sorted(raw_events, key=lambda x: x["time"])
    lots = defaultdict(lambda: {"LONG": deque(), "SHORT": deque()})
    for ev in events:
        key = (ev["account_number"], ev["symbol"])
        qty = float(ev["qty_abs"])
        qty_signed = float(ev["qty_signed"])
        t = pd.to_datetime(ev["time"])
        eff = ev["position_effect"]
        if eff == "OPENING":
            side = "LONG" if qty_signed > 0 else "SHORT"
            lots[key][side].append({"qty": qty, "open_time": t})
            continue
        close_side = "SHORT" if qty_signed > 0 else "LONG"
        qleft = qty
        qbook = lots[key][close_side]
        while qleft > 1e-9 and qbook:
            lot = qbook[0]
            used = min(qleft, float(lot["qty"]))
            lot["qty"] = float(lot["qty"]) - used
            qleft -= used
            if lot["qty"] <= 1e-9:
                qbook.popleft()

    out_rows: List[Dict[str, Any]] = []
    for (acct, sym), side_books in lots.items():
        for side in ("LONG", "SHORT"):
            book = side_books[side]
            if not book:
                continue
            qtys = [float(x.get("qty", 0.0)) for x in book if float(x.get("qty", 0.0)) > 1e-9]
            if not qtys:
                continue
            ts = [
                pd.to_datetime(x.get("open_time"), errors="coerce")
                for x in book
                if float(x.get("qty", 0.0)) > 1e-9
            ]
            qty_ts_pairs = [(q, t) for q, t in zip(qtys, ts) if pd.notna(t)]
            if qty_ts_pairs:
                total_q = sum(q for q, _ in qty_ts_pairs)
                weighted_epoch = sum(q * t.timestamp() for q, t in qty_ts_pairs) / max(1e-9, total_q)
                open_date_est = pd.to_datetime(weighted_epoch, unit="s", utc=True).tz_convert(None)
                _latest_t = max(t for _, t in qty_ts_pairs)
                last_open_trade_time = _latest_t.tz_convert(None) if getattr(_latest_t, "tzinfo", None) else _latest_t
            else:
                open_date_est = pd.NaT
                last_open_trade_time = pd.NaT
            out_rows.append(
                {
                    "account_number": str(acct),
                    "symbol": str(sym).upper(),
                    "side": side,
                    "open_qty_est": float(sum(qtys)),
                    "open_date_est": str(open_date_est) if pd.notna(open_date_est) else "",
                    "last_open_trade_time": str(last_open_trade_time) if pd.notna(last_open_trade_time) else "",
                }
            )
    return pd.DataFrame(out_rows)


def infer_strategy_from_text(
    strategy: str,
    description: str,
    side: str,
    symbol: str = "",
    asset_type: str = "",
    net_quantity: float = math.nan,
) -> str:
    s = f"{strategy} {description} {side} {symbol} {asset_type}".upper()
    _, right, _ = parse_option_meta(symbol)
    has_put = ("PUT" in s) or bool(re.search(r"\bP\b", s)) or (right == "P")
    has_call = ("CALL" in s) or bool(re.search(r"\bC\b", s)) or (right == "C")
    is_short = ("SHORT" in s) or ("SELL" in s) or ("WRITE" in s)
    is_long = ("LONG" in s) or ("BUY" in s)
    if np.isfinite(net_quantity):
        if float(net_quantity) < 0:
            is_short = True
        elif float(net_quantity) > 0:
            is_long = True
    is_option = ("OPTION" in s) or (right in {"C", "P"}) or has_put or has_call
    if is_short and has_put:
        return "Short Put Option"
    if is_short and has_call:
        return "Short Call Option"
    if is_long and has_put:
        return "Long Put Option"
    if is_long and has_call:
        return "Long Call Option"
    if is_option and is_short:
        return "Short Option"
    if is_option and is_long:
        return "Long Option"
    if "EQUITY" in s and is_long:
        return "Long Stock"
    if "EQUITY" in s and is_short:
        return "Short Stock"
    if is_short:
        return "Short"
    if is_long:
        return "Long"
    return "Unknown"


def load_open_positions(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, low_memory=False)
    cols = list(df.columns)
    symbol_col = find_col(cols, ["symbol", "ticker", "underlying"])
    strategy_col = find_col(cols, ["strategy", "strategy_type", "position_type"])
    side_col = find_col(cols, ["side", "action", "direction"])
    desc_col = find_col(cols, ["description", "instrument", "name"])
    expiry_col = find_col(cols, ["expiry", "expiration", "exp_date", "expiration_date"])
    open_date_col = find_col(cols, ["open_date", "open_date_est", "opened_at", "entry_date", "last_open_trade_time"])
    dte_col = find_col(cols, ["dte", "days_to_expiry", "days_to_expiration", "days_left"])
    unreal_col = find_col(cols, ["unrealized_pnl", "unrealized_pl", "open_pnl", "position_pnl"])
    day_pl_col = find_col(cols, ["current_day_profit_loss", "day_pnl", "daily_pnl"])
    avg_price_col = find_col(cols, ["average_price", "avg_price", "average_cost", "avg_cost"])
    market_value_col = find_col(cols, ["market_value", "position_value", "marketvalue"])
    long_qty_col = find_col(cols, ["long_quantity", "long_qty", "quantity_long"])
    short_qty_col = find_col(cols, ["short_quantity", "short_qty", "quantity_short"])
    qty_col = find_col(cols, ["quantity", "qty", "net_quantity"])
    asset_type_col = find_col(cols, ["asset_type", "instrument_type", "security_type"])
    max_profit_col = find_col(cols, ["max_profit", "profit_cap", "target_profit"])
    risk_col = find_col(
        cols,
        [
            "max_loss",
            "risk",
            "risk_amount",
            "max_risk",
            "maintenance_requirement",
            "buying_power_effect",
            "bp_effect",
            "margin_requirement",
        ],
    )

    out = pd.DataFrame(index=df.index)
    out["symbol"] = df[symbol_col].astype(str).str.strip().str.upper() if symbol_col else "UNKNOWN"
    out["strategy_raw"] = df[strategy_col].astype(str).str.strip() if strategy_col else ""
    out["description"] = df[desc_col].astype(str).str.strip() if desc_col else ""
    out["side_raw"] = df[side_col].astype(str).str.strip() if side_col else ""
    out["expiry"] = pd.to_datetime(df[expiry_col], errors="coerce") if expiry_col else pd.NaT
    out["open_date"] = pd.to_datetime(df[open_date_col], errors="coerce") if open_date_col else pd.NaT
    out["dte"] = df[dte_col].map(parse_float) if dte_col else math.nan
    out["asset_type"] = df[asset_type_col].astype(str).str.upper().str.strip() if asset_type_col else ""
    out["long_quantity"] = df[long_qty_col].map(parse_float).fillna(0.0) if long_qty_col else 0.0
    out["short_quantity"] = df[short_qty_col].map(parse_float).fillna(0.0) if short_qty_col else 0.0
    out["raw_quantity"] = df[qty_col].map(parse_float) if qty_col else math.nan
    out["net_quantity"] = out["raw_quantity"]
    nq_fallback = out["long_quantity"] - out["short_quantity"]
    out.loc[out["net_quantity"].isna(), "net_quantity"] = nq_fallback[out["net_quantity"].isna()]
    out["average_price"] = df[avg_price_col].map(parse_float) if avg_price_col else math.nan
    out["market_value"] = df[market_value_col].map(parse_float) if market_value_col else math.nan
    out["unrealized_pnl"] = df[unreal_col].map(parse_float) if unreal_col else math.nan
    out["current_day_profit_loss"] = df[day_pl_col].map(parse_float) if day_pl_col else math.nan
    out["max_profit"] = df[max_profit_col].map(parse_float).abs() if max_profit_col else math.nan
    out["risk"] = df[risk_col].map(parse_float).abs() if risk_col else math.nan
    out["underlying"] = out["symbol"].map(option_underlying_symbol)
    out.loc[out["underlying"].isna() | (out["underlying"] == ""), "underlying"] = out.loc[
        out["underlying"].isna() | (out["underlying"] == ""), "symbol"
    ].astype(str)
    out["strategy"] = [
        infer_strategy_from_text(a, b, c, d, e, f)
        for a, b, c, d, e, f in zip(
            out["strategy_raw"],
            out["description"],
            out["side_raw"],
            out["symbol"],
            out["asset_type"],
            out["net_quantity"],
        )
    ]
    out["symbol"] = out["symbol"].replace({"": "UNKNOWN", "NAN": "UNKNOWN"})

    # Estimate unrealized P/L when not explicitly present.
    mult = np.where(out["asset_type"].str.contains("OPTION", na=False), 100.0, 1.0)
    abs_qty = out["net_quantity"].abs()
    cost_basis_abs = out["average_price"] * abs_qty * mult
    can_est = out["unrealized_pnl"].isna() & out["average_price"].notna() & out["market_value"].notna() & abs_qty.gt(0)
    long_mask = can_est & out["net_quantity"].gt(0)
    short_mask = can_est & out["net_quantity"].lt(0)
    out.loc[long_mask, "unrealized_pnl"] = out.loc[long_mask, "market_value"] - cost_basis_abs[long_mask]
    out.loc[short_mask, "unrealized_pnl"] = cost_basis_abs[short_mask] + out.loc[short_mask, "market_value"]
    out["unrealized_source"] = np.where(
        out["unrealized_pnl"].notna(),
        np.where(df[unreal_col].map(parse_float).notna() if unreal_col else False, "explicit", "estimated"),
        np.where(out["current_day_profit_loss"].notna(), "day_only", "missing"),
    )
    out.loc[out["unrealized_pnl"].isna() & out["current_day_profit_loss"].notna(), "unrealized_pnl"] = out.loc[
        out["unrealized_pnl"].isna() & out["current_day_profit_loss"].notna(), "current_day_profit_loss"
    ]
    return out


def parse_option_meta(symbol: str) -> Tuple[Optional[dt.date], Optional[str], Optional[float]]:
    text = str(symbol or "").upper()
    m = SCHWAB_OPTION_SYMBOL_RE.match(text)
    if m:
        _, yymmdd, right, strike8 = m.groups()
    else:
        m2 = COMPACT_OPTION_SYMBOL_RE.match(text.replace(" ", ""))
        if not m2:
            return None, None, None
        _, yymmdd, right, strike8 = m2.groups()
    try:
        expiry = dt.datetime.strptime(yymmdd, "%y%m%d").date()
    except Exception:
        expiry = None
    try:
        strike = float(int(strike8)) / 1000.0
    except Exception:
        strike = None
    return expiry, right, strike


def option_underlying_symbol(symbol: str) -> str:
    text = str(symbol or "").upper()
    m = SCHWAB_OPTION_SYMBOL_RE.match(text)
    if m:
        return m.group(1).strip()
    m2 = COMPACT_OPTION_SYMBOL_RE.match(text.replace(" ", ""))
    if m2:
        return m2.group(1).strip()
    return ""


def _alloc_pro_rata(value: Any, part_qty: float, total_qty: float) -> float:
    v = parse_float(value)
    q_part = parse_float(part_qty)
    q_total = parse_float(total_qty)
    if not np.isfinite(v) or not np.isfinite(q_part) or not np.isfinite(q_total) or abs(q_total) <= 1e-9:
        return math.nan
    return float(v) * (abs(float(q_part)) / abs(float(q_total)))


def _flatten_chain_contracts(chain_payload: Dict[str, Any], right: str) -> List[Dict[str, Any]]:
    if not isinstance(chain_payload, dict):
        return []
    right_u = str(right or "").upper()
    key = "putExpDateMap" if right_u == "P" else "callExpDateMap"
    exp_map = chain_payload.get(key, {}) if isinstance(chain_payload, dict) else {}
    out: List[Dict[str, Any]] = []
    if not isinstance(exp_map, dict):
        return out
    for exp_key, strike_map in exp_map.items():
        if not isinstance(strike_map, dict):
            continue
        exp_date_txt = str(exp_key).split(":")[0].strip()
        exp_date = pd.to_datetime(exp_date_txt, errors="coerce")
        for strike_key, contracts in strike_map.items():
            strike = parse_float(strike_key)
            if not isinstance(contracts, list):
                continue
            for c in contracts:
                if not isinstance(c, dict):
                    continue
                out.append(
                    {
                        "symbol": str(c.get("symbol", "")).strip().upper(),
                        "expiry": exp_date.date() if pd.notna(exp_date) else None,
                        "strike": strike,
                        "bid": parse_float(c.get("bid")),
                        "ask": parse_float(c.get("ask")),
                        "mark": parse_float(c.get("mark")),
                    }
                )
    return out


def _quote_last_price(quotes_payload: Dict[str, Any], underlying: str) -> float:
    q = (quotes_payload or {}).get(str(underlying).upper(), {})
    body = q.get("quote", q) if isinstance(q, dict) else {}
    return parse_float(body.get("lastPrice", body.get("mark")))


def _find_contract_quote(chain_payload: Dict[str, Any], option_symbol: str) -> Dict[str, Any]:
    sym = str(option_symbol or "").strip().upper()
    if not sym:
        return {}
    for right in ("P", "C"):
        for c in _flatten_chain_contracts(chain_payload, right):
            if str(c.get("symbol", "")).upper() == sym:
                return c
    return {}


def _estimate_roll_plan(
    structure: Dict[str, Any],
    as_of_date: dt.date,
    cfg: Dict,
    live_context: Optional[Dict[str, Any]],
) -> str:
    if not isinstance(live_context, dict):
        return ""
    underlying = str(structure.get("underlying", "")).strip().upper()
    right = str(structure.get("right", "")).strip().upper()
    short_sym = str(structure.get("short_leg_symbol", "")).strip().upper()
    long_sym = str(structure.get("long_leg_symbol", "")).strip().upper()
    if (not underlying) or (right not in {"P", "C"}) or (not short_sym):
        return ""

    chains = live_context.get("option_chains", {})
    quotes = live_context.get("quotes", {})
    chain_payload = chains.get(underlying, {}) if isinstance(chains, dict) else {}
    if not isinstance(chain_payload, dict) or not chain_payload:
        return "Need live chain for roll targets."

    pm = cfg.get("playbook", {}).get("position_management", {}
    )
    min_extend = int(pm.get("roll_extend_min_days", 14))
    max_extend = int(pm.get("roll_extend_max_days", 35))
    otm_pct = float(pm.get("roll_target_otm_pct", 0.08))
    target_mid = min_extend + max(1, (max_extend - min_extend) // 2)

    spot = _quote_last_price(quotes, underlying)
    contracts = _flatten_chain_contracts(chain_payload, right)
    if not contracts:
        return "Need live option chain contracts for roll targets."

    cur_short = _find_contract_quote(chain_payload, short_sym)
    cur_short_ask = parse_float(cur_short.get("ask"))
    cur_short_bid = parse_float(cur_short.get("bid"))
    cur_short_mid = np.nanmean([cur_short_bid, cur_short_ask]) if np.isfinite(cur_short_bid) or np.isfinite(cur_short_ask) else math.nan

    dte_now = parse_float(structure.get("dte", math.nan))
    min_new_dte = int(dte_now + min_extend) if np.isfinite(dte_now) else min_extend
    max_new_dte = int(dte_now + max_extend) if np.isfinite(dte_now) else (max_extend + 120)

    candidates: List[Dict[str, Any]] = []
    for c in contracts:
        exp = c.get("expiry")
        strike = parse_float(c.get("strike"))
        bid = parse_float(c.get("bid"))
        ask = parse_float(c.get("ask"))
        if exp is None or not np.isfinite(strike):
            continue
        dte_new = (exp - as_of_date).days
        if dte_new < min_new_dte or dte_new > max_new_dte:
            continue
        if right == "P":
            desired = (spot * (1.0 - otm_pct)) if np.isfinite(spot) else strike
            if np.isfinite(spot) and strike > desired:
                continue
            dist = abs(strike - desired) if np.isfinite(desired) else abs(strike)
        else:
            desired = (spot * (1.0 + otm_pct)) if np.isfinite(spot) else strike
            if np.isfinite(spot) and strike < desired:
                continue
            dist = abs(strike - desired) if np.isfinite(desired) else abs(strike)
        dte_penalty = abs(dte_new - (int(dte_now) + target_mid)) if np.isfinite(dte_now) else abs(dte_new - target_mid)
        candidates.append(
            {
                "symbol": str(c.get("symbol", "")).strip().upper(),
                "expiry": exp,
                "strike": strike,
                "bid": bid,
                "ask": ask,
                "rank": (dte_penalty, dist),
            }
        )

    if not candidates:
        return "No candidate roll strikes in target DTE window."
    tgt = sorted(candidates, key=lambda x: x["rank"])[0]
    new_short_bid = parse_float(tgt.get("bid"))
    new_short_ask = parse_float(tgt.get("ask"))
    new_short_mid = np.nanmean([new_short_bid, new_short_ask]) if np.isfinite(new_short_bid) or np.isfinite(new_short_ask) else math.nan

    is_spread = bool(long_sym)
    if not is_spread:
        if np.isfinite(new_short_bid) and np.isfinite(cur_short_ask):
            roll_net = new_short_bid - cur_short_ask
            net_txt = f"{roll_net:+.2f} cr/db est"
        elif np.isfinite(new_short_mid) and np.isfinite(cur_short_mid):
            roll_net = new_short_mid - cur_short_mid
            net_txt = f"{roll_net:+.2f} cr/db est(mid)"
        else:
            net_txt = "net n/a"
        return (
            f"Roll short {right} to {tgt['strike']:.2f} {right} @ {tgt['expiry']} "
            f"(bid/ask {new_short_bid if np.isfinite(new_short_bid) else 'n/a'}/"
            f"{new_short_ask if np.isfinite(new_short_ask) else 'n/a'}); est roll {net_txt}."
        )

    cur_long = _find_contract_quote(chain_payload, long_sym)
    cur_long_bid = parse_float(cur_long.get("bid"))
    cur_long_ask = parse_float(cur_long.get("ask"))
    cur_long_mid = np.nanmean([cur_long_bid, cur_long_ask]) if np.isfinite(cur_long_bid) or np.isfinite(cur_long_ask) else math.nan
    short_strike = parse_float(structure.get("short_strike", math.nan))
    long_strike = parse_float(structure.get("long_strike", math.nan))
    width = abs(short_strike - long_strike) if np.isfinite(short_strike) and np.isfinite(long_strike) else parse_float(structure.get("width", math.nan))
    if not np.isfinite(width) or width <= 0:
        width = 5.0
    new_long_strike = (tgt["strike"] - width) if right == "P" else (tgt["strike"] + width)
    same_exp = [c for c in contracts if c.get("expiry") == tgt["expiry"] and np.isfinite(parse_float(c.get("strike")))]
    if not same_exp:
        return f"Roll short to {tgt['strike']:.2f}{right} @ {tgt['expiry']}; no hedge leg found."
    new_long_candidates = sorted(same_exp, key=lambda c: abs(parse_float(c.get("strike")) - new_long_strike))
    new_long = new_long_candidates[0]
    new_long_ask = parse_float(new_long.get("ask"))
    new_long_bid = parse_float(new_long.get("bid"))
    new_long_mid = np.nanmean([new_long_bid, new_long_ask]) if np.isfinite(new_long_bid) or np.isfinite(new_long_ask) else math.nan

    if np.isfinite(new_short_bid) and np.isfinite(new_long_ask) and np.isfinite(cur_short_ask) and np.isfinite(cur_long_bid):
        roll_net = (new_short_bid - new_long_ask) - (cur_short_ask - cur_long_bid)
        net_txt = f"{roll_net:+.2f} cr/db est"
    elif np.isfinite(new_short_mid) and np.isfinite(new_long_mid) and np.isfinite(cur_short_mid) and np.isfinite(cur_long_mid):
        roll_net = (new_short_mid - new_long_mid) - (cur_short_mid - cur_long_mid)
        net_txt = f"{roll_net:+.2f} cr/db est(mid)"
    else:
        net_txt = "net n/a"
    return (
        f"Roll to {tgt['strike']:.2f}{right}/{parse_float(new_long.get('strike')):.2f}{right} @ {tgt['expiry']} "
        f"(same width {width:.2f}); est roll {net_txt}."
    )


def build_position_decisions(
    pos: pd.DataFrame,
    as_of_date: dt.date,
    cfg: Dict,
    live_context: Optional[Dict[str, Any]] = None,
) -> pd.DataFrame:
    if pos.empty:
        return pd.DataFrame()

    pm = cfg.get("playbook", {}).get("position_management", {})
    take_profit_default = float(pm.get("take_profit_pct_of_max_profit", 0.60))
    stop_loss_default = float(pm.get("stop_loss_pct_of_risk", 0.50))
    take_profit_credit = float(pm.get("take_profit_pct_credit_max_profit", take_profit_default))
    take_profit_debit = float(pm.get("take_profit_pct_debit_cost", max(take_profit_default, 0.80)))
    stop_loss_credit = float(pm.get("stop_loss_pct_credit_max_loss", max(stop_loss_default, 0.65)))
    stop_loss_debit = float(pm.get("stop_loss_pct_debit_max_loss", max(stop_loss_default, 0.55)))
    short_roll_min_dte = int(pm.get("short_roll_min_dte", 7))
    short_roll_max_dte = int(pm.get("short_roll_max_dte", max(int(pm.get("short_roll_dte_threshold", 10)), 21)))
    long_exit_dte = int(pm.get("long_option_exit_dte_threshold", 14))
    near_expiry_dte = int(pm.get("near_expiry_dte_threshold", 5))
    roll_risk_draw_trigger = float(pm.get("roll_risk_draw_trigger", 0.20))

    p = pos.copy().reset_index(drop=True)
    p["row_id"] = np.arange(len(p), dtype=int)
    parsed = p["symbol"].map(parse_option_meta)
    p["opt_expiry_sym"] = [x[0] for x in parsed]
    p["option_right"] = [x[1] for x in parsed]
    p["option_strike"] = [x[2] for x in parsed]
    p["expiry"] = pd.to_datetime(p["expiry"], errors="coerce")
    missing_exp = p["expiry"].isna() & p["opt_expiry_sym"].notna()
    p.loc[missing_exp, "expiry"] = pd.to_datetime(p.loc[missing_exp, "opt_expiry_sym"], errors="coerce")
    p["dte"] = p["dte"].map(parse_float)
    missing_dte = p["dte"].isna() & p["expiry"].notna()
    p.loc[missing_dte, "dte"] = (p.loc[missing_dte, "expiry"].dt.date - as_of_date).map(
        lambda x: float(x.days) if pd.notna(x) else math.nan
    )
    p["net_quantity"] = p["net_quantity"].map(parse_float)
    p["abs_qty"] = p["net_quantity"].abs()
    p = p[p["abs_qty"].fillna(0) > 0].copy()
    if p.empty:
        return pd.DataFrame()

    side_from_strategy = p["strategy"].astype(str).str.contains("Short", case=False, na=False).map(
        lambda is_short: -1.0 if is_short else 1.0
    )
    p["side_sign"] = np.where(p["net_quantity"].notna() & (p["net_quantity"] != 0), np.sign(p["net_quantity"]), side_from_strategy)
    p["side_sign"] = p["side_sign"].replace(0, 1)
    p["underlying"] = p["underlying"].astype(str).str.strip().str.upper()
    p["underlying"] = np.where(p["underlying"] == "", p["symbol"].astype(str).str.upper(), p["underlying"])

    prow = p.set_index("row_id", drop=False)
    remaining: Dict[int, float] = {int(r.row_id): float(r.abs_qty) for r in p.itertuples(index=False)}
    structures: List[Dict[str, Any]] = []
    sid_counter = 1

    option_rows = p[p["option_right"].isin(["C", "P"])].copy()
    option_rows["expiry_date"] = option_rows["expiry"].dt.date
    for (underlying, expiry_date, right), g in option_rows.groupby(["underlying", "expiry_date", "option_right"], dropna=False):
        if not isinstance(underlying, str) or not underlying:
            continue
        longs = [int(x) for x in g[g["side_sign"] > 0]["row_id"].tolist()]
        shorts = [int(x) for x in g[g["side_sign"] < 0]["row_id"].tolist()]

        while True:
            candidates: List[Tuple[float, float, int, int]] = []
            for sid in shorts:
                if remaining.get(sid, 0.0) <= 1e-9:
                    continue
                srow = prow.loc[sid]
                s_strike = parse_float(srow["option_strike"])
                s_open = pd.to_datetime(srow.get("open_date"), errors="coerce")
                for lid in longs:
                    if remaining.get(lid, 0.0) <= 1e-9:
                        continue
                    lrow = prow.loc[lid]
                    l_strike = parse_float(lrow["option_strike"])
                    l_open = pd.to_datetime(lrow.get("open_date"), errors="coerce")
                    width = abs(s_strike - l_strike) if np.isfinite(s_strike) and np.isfinite(l_strike) else 1e9
                    if pd.notna(s_open) and pd.notna(l_open):
                        open_gap_days = abs((s_open - l_open).total_seconds()) / 86400.0
                    else:
                        open_gap_days = 1e9
                    candidates.append((open_gap_days, width, sid, lid))
            if not candidates:
                break
            _, _, sid, lid = sorted(candidates, key=lambda x: (x[0], x[1]))[0]
            pair_qty = min(float(remaining.get(sid, 0.0)), float(remaining.get(lid, 0.0)))
            if pair_qty <= 1e-9:
                break
            srow = prow.loc[sid]
            lrow = prow.loc[lid]
            short_strike = parse_float(srow["option_strike"])
            long_strike = parse_float(lrow["option_strike"])
            short_avg = parse_float(srow["average_price"])
            long_avg = parse_float(lrow["average_price"])
            entry_net_cash = (
                (short_avg - long_avg) * 100.0 * pair_qty
                if np.isfinite(short_avg) and np.isfinite(long_avg)
                else math.nan
            )
            orient_net = "credit" if (np.isfinite(entry_net_cash) and entry_net_cash >= 0) else "debit"
            if right == "P":
                orient_shape = "credit" if (
                    np.isfinite(short_strike) and np.isfinite(long_strike) and short_strike > long_strike
                ) else "debit"
            else:
                orient_shape = "credit" if (
                    np.isfinite(short_strike) and np.isfinite(long_strike) and short_strike < long_strike
                ) else "debit"
            net_type = orient_net if np.isfinite(entry_net_cash) else orient_shape
            if right == "P":
                strategy = "Bull Put Credit" if net_type == "credit" else "Bear Put Debit"
            else:
                strategy = "Bear Call Credit" if net_type == "credit" else "Bull Call Debit"

            width_cash = (
                abs(short_strike - long_strike) * 100.0 * pair_qty
                if np.isfinite(short_strike) and np.isfinite(long_strike)
                else math.nan
            )
            risk_fallback = np.nansum(
                [
                    _alloc_pro_rata(srow["risk"], pair_qty, srow["abs_qty"]),
                    _alloc_pro_rata(lrow["risk"], pair_qty, lrow["abs_qty"]),
                ]
            )
            if net_type == "credit":
                max_profit = max(entry_net_cash, 0.0) if np.isfinite(entry_net_cash) else math.nan
                max_loss = (width_cash - max_profit) if np.isfinite(width_cash) and np.isfinite(max_profit) else math.nan
            else:
                debit_cash = abs(entry_net_cash) if np.isfinite(entry_net_cash) else math.nan
                max_loss = debit_cash
                max_profit = (width_cash - debit_cash) if np.isfinite(width_cash) and np.isfinite(debit_cash) else math.nan
            if (not np.isfinite(max_loss)) or max_loss <= 0:
                max_loss = risk_fallback if np.isfinite(risk_fallback) and risk_fallback > 0 else math.nan
            if np.isfinite(max_profit) and max_profit < 0:
                max_profit = math.nan

            unreal = np.nansum(
                [
                    _alloc_pro_rata(srow["unrealized_pnl"], pair_qty, srow["abs_qty"]),
                    _alloc_pro_rata(lrow["unrealized_pnl"], pair_qty, lrow["abs_qty"]),
                ]
            )
            dte_vals = [parse_float(srow["dte"]), parse_float(lrow["dte"])]
            dte_vals = [x for x in dte_vals if np.isfinite(x)]
            dte = min(dte_vals) if dte_vals else math.nan
            expiry_txt = (
                str(expiry_date)
                if pd.notna(expiry_date)
                else str(pd.to_datetime(srow["expiry"]).date()) if pd.notna(srow["expiry"]) else ""
            )
            position_id = f"P{sid_counter:03d}-{underlying}-{right}-{expiry_txt}"
            sid_counter += 1
            structures.append(
                {
                    "position_id": position_id,
                    "underlying": underlying,
                    "symbol": position_id,
                    "strategy": strategy,
                    "legs": f"{str(srow['symbol']).strip()} (short) + {str(lrow['symbol']).strip()} (long)",
                    "expiry": expiry_txt,
                    "dte": int(dte) if np.isfinite(dte) else math.nan,
                    "net_type": net_type,
                    "entry_cost": abs(entry_net_cash) if np.isfinite(entry_net_cash) else math.nan,
                    "unrealized_pnl": unreal if np.isfinite(unreal) else math.nan,
                    "max_profit": max_profit if np.isfinite(max_profit) else math.nan,
                    "risk": max_loss if np.isfinite(max_loss) else math.nan,
                    "unrealized_source": "estimated",
                    "right": right,
                    "short_leg_symbol": str(srow["symbol"]).strip().upper(),
                    "long_leg_symbol": str(lrow["symbol"]).strip().upper(),
                    "short_strike": short_strike,
                    "long_strike": long_strike,
                    "width": abs(short_strike - long_strike) if np.isfinite(short_strike) and np.isfinite(long_strike) else math.nan,
                }
            )
            remaining[sid] = max(0.0, float(remaining.get(sid, 0.0)) - pair_qty)
            remaining[lid] = max(0.0, float(remaining.get(lid, 0.0)) - pair_qty)

    for rid, rem_qty in remaining.items():
        if rem_qty <= 1e-9:
            continue
        r = prow.loc[rid]
        sym = str(r["symbol"]).strip().upper() or "UNKNOWN"
        right = str(r["option_right"] or "").upper()
        side_sign = parse_float(r["side_sign"])
        is_short = np.isfinite(side_sign) and side_sign < 0
        expiry = pd.to_datetime(r["expiry"], errors="coerce")
        dte = parse_float(r["dte"])
        if (not np.isfinite(dte)) and pd.notna(expiry):
            dte = float((expiry.date() - as_of_date).days)
        avg = parse_float(r["average_price"])
        entry_cash = avg * 100.0 * rem_qty if np.isfinite(avg) else math.nan
        unreal = _alloc_pro_rata(r["unrealized_pnl"], rem_qty, r["abs_qty"])
        risk_alloc = _alloc_pro_rata(r["risk"], rem_qty, r["abs_qty"])
        strike = parse_float(r["option_strike"])
        if is_short:
            strategy = "Short Put Option" if right == "P" else "Short Call Option"
            net_type = "credit"
            max_profit = entry_cash if np.isfinite(entry_cash) else math.nan
            max_loss = risk_alloc if np.isfinite(risk_alloc) and risk_alloc > 0 else math.nan
        else:
            strategy = "Long Put Option" if right == "P" else "Long Call Option"
            net_type = "debit"
            max_loss = risk_alloc if np.isfinite(risk_alloc) and risk_alloc > 0 else entry_cash
            if right == "P" and np.isfinite(strike) and np.isfinite(entry_cash):
                max_profit = max(0.0, (strike * 100.0 * rem_qty) - entry_cash)
            else:
                max_profit = math.nan
        position_id = f"P{sid_counter:03d}-{str(r['underlying']).strip().upper()}-{right}-{str(expiry.date()) if pd.notna(expiry) else ''}-{int(strike) if np.isfinite(strike) else 'NA'}-{('S' if is_short else 'L')}"
        sid_counter += 1
        structures.append(
            {
                "position_id": position_id,
                "underlying": str(r["underlying"]).strip().upper() or sym,
                "symbol": position_id,
                "strategy": strategy,
                "legs": f"{sym} x{rem_qty:g}",
                "expiry": str(expiry.date()) if pd.notna(expiry) else "",
                "dte": int(dte) if np.isfinite(dte) else math.nan,
                "net_type": net_type,
                "entry_cost": abs(entry_cash) if np.isfinite(entry_cash) else math.nan,
                "unrealized_pnl": unreal if np.isfinite(unreal) else math.nan,
                "max_profit": max_profit if np.isfinite(max_profit) else math.nan,
                "risk": max_loss if np.isfinite(max_loss) else math.nan,
                "unrealized_source": str(r.get("unrealized_source", "missing") if isinstance(r, pd.Series) else "estimated"),
                "right": right,
                "short_leg_symbol": sym if is_short else "",
                "long_leg_symbol": "" if is_short else sym,
                "short_strike": strike if is_short else math.nan,
                "long_strike": strike if (not is_short) else math.nan,
                "width": math.nan,
            }
        )

    rows: List[Dict[str, Any]] = []
    for s in structures:
        dte = parse_float(s.get("dte", math.nan))
        unreal = parse_float(s.get("unrealized_pnl", math.nan))
        max_profit = parse_float(s.get("max_profit", math.nan))
        risk = parse_float(s.get("risk", math.nan))
        entry_cost = parse_float(s.get("entry_cost", math.nan))
        net_type = str(s.get("net_type", "")).strip().lower()
        strategy = str(s.get("strategy", "Unknown")).strip() or "Unknown"
        unreal_source = str(s.get("unrealized_source", "missing")).strip() or "missing"

        profit_capture = (unreal / max_profit) if (np.isfinite(unreal) and np.isfinite(max_profit) and max_profit > 0) else math.nan
        pnl_vs_cost = (unreal / entry_cost) if (np.isfinite(unreal) and np.isfinite(entry_cost) and entry_cost > 0) else math.nan
        risk_draw = ((-unreal) / risk) if (np.isfinite(unreal) and np.isfinite(risk) and risk > 0 and unreal < 0) else 0.0

        rec = "WAIT"
        reason = "Inside tolerance/no hard trigger; keep monitoring."
        priority = 8

        if unreal_source == "day_only":
            rec = "DATA_NEEDED"
            reason = "Only day P/L available; need full unrealized P/L for reliable close/roll."
            priority = 0
        elif not np.isfinite(unreal):
            rec = "DATA_NEEDED"
            reason = "Missing unrealized P/L; cannot score recoverability."
            priority = 0
        else:
            # Debit stop is DTE-aware; farther-dated debits get wider breathing room.
            debit_stop_near = float(pm.get("stop_loss_pct_debit_near_dte", 0.45))
            debit_stop_mid = float(pm.get("stop_loss_pct_debit_mid_dte", stop_loss_debit))
            debit_stop_far = float(pm.get("stop_loss_pct_debit_far_dte", 0.65))
            debit_mid_min_dte = int(pm.get("stop_loss_pct_debit_mid_min_dte", 21))
            debit_far_min_dte = int(pm.get("stop_loss_pct_debit_far_min_dte", 45))
            leaps_dte_threshold = int(pm.get("leaps_dte_threshold", 120))
            leaps_stop_pct = float(pm.get("leaps_debit_catastrophic_stop_pct", 0.75))

            if net_type == "credit":
                if np.isfinite(profit_capture) and profit_capture >= take_profit_credit:
                    rec = "CLOSE_NOW"
                    reason = f"Credit profit capture {profit_capture:.0%} >= {take_profit_credit:.0%} target."
                    priority = 1
                elif np.isfinite(risk_draw) and risk_draw >= stop_loss_credit:
                    rec = "CLOSE_NOW"
                    reason = f"Credit loss at {risk_draw:.0%} of max risk >= {stop_loss_credit:.0%} stop."
                    priority = 1
                elif (
                    np.isfinite(dte)
                    and short_roll_min_dte <= dte <= short_roll_max_dte
                    and np.isfinite(risk_draw)
                    and risk_draw >= roll_risk_draw_trigger
                ):
                    rec = "ROLL"
                    reason = (
                        f"Short premium in roll window ({int(dte)} DTE) with risk draw {risk_draw:.0%}; "
                        "roll out 14-35 days and move short strike away."
                    )
                    priority = 3
            else:
                if np.isfinite(dte) and dte >= leaps_dte_threshold:
                    debit_stop_use = max(leaps_stop_pct, debit_stop_far)
                elif np.isfinite(dte) and dte >= debit_far_min_dte:
                    debit_stop_use = debit_stop_far
                elif np.isfinite(dte) and dte >= debit_mid_min_dte:
                    debit_stop_use = debit_stop_mid
                else:
                    debit_stop_use = debit_stop_near

                if np.isfinite(pnl_vs_cost) and pnl_vs_cost >= take_profit_debit:
                    rec = "CLOSE_NOW"
                    reason = f"Debit position return {pnl_vs_cost:.0%} >= {take_profit_debit:.0%} target."
                    priority = 1
                elif np.isfinite(risk_draw) and risk_draw >= debit_stop_use:
                    rec = "CLOSE_NOW"
                    if np.isfinite(dte) and dte >= leaps_dte_threshold:
                        reason = (
                            f"LEAPS catastrophic stop: loss {risk_draw:.0%} >= {debit_stop_use:.0%} "
                            f"(DTE {int(dte)})."
                        )
                    else:
                        reason = f"Debit loss at {risk_draw:.0%} of max loss >= {debit_stop_use:.0%} DTE-tier stop."
                    priority = 1
                elif strategy in {"Long Call Option", "Long Put Option"} and np.isfinite(dte) and dte <= long_exit_dte and unreal <= 0:
                    rec = "CLOSE_NOW"
                    reason = f"Long option decay risk: DTE {int(dte)} <= {long_exit_dte} and non-positive P/L."
                    priority = 2

            if rec == "WAIT" and np.isfinite(dte) and dte <= near_expiry_dte and unreal <= 0:
                rec = "CLOSE_NOW"
                reason = f"Near expiry (DTE {int(dte)} <= {near_expiry_dte}) with weak reward-to-risk."
                priority = 2

        action = {
            "CLOSE_NOW": "🔴 CLOSE",
            "ROLL": "🟣 ROLL",
            "WAIT": "🟡 WAIT",
            "DATA_NEEDED": "⚪ NEEDS DATA",
        }.get(rec, rec)
        roll_plan = _estimate_roll_plan(s, as_of_date, cfg, live_context) if rec == "ROLL" else ""

        rows.append(
            {
                "position_id": str(s.get("position_id", "")),
                "underlying": str(s.get("underlying", "")).strip().upper(),
                "symbol": str(s.get("symbol", "")),
                "strategy": strategy,
                "legs": str(s.get("legs", "")),
                "expiry": str(s.get("expiry", "")),
                "dte": int(dte) if np.isfinite(dte) else math.nan,
                "unrealized_pnl": unreal if np.isfinite(unreal) else math.nan,
                "risk": risk if np.isfinite(risk) else math.nan,
                "profit_capture_pct": profit_capture if np.isfinite(profit_capture) else math.nan,
                "risk_draw_pct": risk_draw if np.isfinite(risk_draw) else math.nan,
                "recommendation": rec,
                "action": action,
                "reason": reason,
                "roll_plan": roll_plan,
                "unrealized_source": unreal_source,
                "_priority": priority,
            }
        )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out = out.sort_values(["_priority", "underlying", "dte"], ascending=[True, True, True]).reset_index(drop=True)
    return out.drop(columns=["_priority"])


def write_json(path: Path, payload: Dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def to_md_table(df: pd.DataFrame, n: int = 30) -> str:
    if df.empty:
        return "_none_"
    return df.head(n).to_markdown(index=False)


def run_daily(
    trades: pd.DataFrame,
    out_dir: Path,
    cfg: Dict,
    open_positions_csv: Optional[Path],
    lookback_trades: int,
) -> Dict:
    pb = cfg.get("playbook", {})
    daily_cfg = pb.get("daily", {})
    risk_cfg = pb.get("risk_limits", {})

    daily_loss_stop_cash = float(daily_cfg.get("daily_loss_stop_cash", 1000.0))
    streak_yellow = int(daily_cfg.get("loss_streak_yellow", 3))
    streak_red = int(daily_cfg.get("loss_streak_red", 4))
    min_pf_yellow = float(daily_cfg.get("rolling_pf_yellow_floor", 1.0))
    short_put_limit = float(risk_cfg.get("short_put_max_share", 0.35))
    symbol_limit = float(risk_cfg.get("single_symbol_max_share", 0.10))
    expiry_limit = float(risk_cfg.get("single_expiry_max_share_short_put", 0.25))

    x = trades.sort_values("date").reset_index(drop=True)
    recent = x.tail(max(1, int(lookback_trades))).copy()
    last_day = x["date"].dt.date.max()
    day_df = x[x["date"].dt.date == last_day].copy()

    day_pnl = float(day_df["realized_pnl"].sum()) if not day_df.empty else 0.0
    recent_pf = compute_profit_factor(recent["realized_pnl"])
    recent_net = float(recent["realized_pnl"].sum())
    cur_loss_streak = trailing_loss_streak(x["realized_pnl"])

    status = "GREEN"
    severity = 0
    triggers: List[str] = []
    actions: List[str] = []

    def escalate(level: str, trigger: str, action: str) -> None:
        nonlocal status, severity
        rank = {"GREEN": 0, "YELLOW": 1, "RED": 2}[level]
        if rank > severity:
            severity = rank
            status = level
        triggers.append(trigger)
        actions.append(action)

    if day_pnl <= -abs(daily_loss_stop_cash):
        escalate(
            "RED",
            f"Daily realized P/L {day_pnl:,.2f} <= stop {-abs(daily_loss_stop_cash):,.2f}",
            "Stop opening new risk today; only reduce or hedge open risk.",
        )
    if cur_loss_streak >= streak_red:
        escalate(
            "RED",
            f"Trailing loss streak is {cur_loss_streak} (red threshold {streak_red})",
            "Cut size by 50% and skip the next 2 non-core setups.",
        )
    elif cur_loss_streak >= streak_yellow:
        escalate(
            "YELLOW",
            f"Trailing loss streak is {cur_loss_streak} (yellow threshold {streak_yellow})",
            "Reduce new position size by 25% until streak resets.",
        )
    if np.isfinite(recent_pf) and recent_pf < min_pf_yellow:
        escalate(
            "YELLOW",
            f"Recent {len(recent)}-trade PF is {recent_pf:.2f} < {min_pf_yellow:.2f}",
            "Restrict entries to top-conviction setups only this week.",
        )

    pos_summary: Dict[str, float] = {}
    symbol_risk_table = pd.DataFrame()
    expiry_risk_table = pd.DataFrame()
    position_decisions = pd.DataFrame()
    if open_positions_csv and open_positions_csv.exists():
        pos = load_open_positions(open_positions_csv)
        options_only = bool(daily_cfg.get("options_only", True))
        if options_only:
            is_option = (
                pos["asset_type"].astype(str).str.upper().eq("OPTION")
                | pos["strategy"].astype(str).str.contains("Option", case=False, na=False)
                | pos["symbol"].astype(str).str.contains(r"\d{6}[CP]\d{8}", na=False)
            )
            pos = pos[is_option].copy()
        live_context: Dict[str, Any] = {}
        pm = cfg.get("playbook", {}).get("position_management", {})
        if bool(pm.get("roll_use_live_chain", True)):
            try:
                from uwos.schwab_auth import SchwabAuthConfig, SchwabLiveDataService

                chain_symbols = sorted(
                    {
                        str(x).strip().upper()
                        for x in pos["underlying"].dropna().tolist()
                        if str(x).strip()
                    }
                )
                if chain_symbols:
                    cfg_live = SchwabAuthConfig.from_env(load_dotenv_file=True)
                    svc_live = SchwabLiveDataService(cfg_live)
                    svc_live.connect()
                    roll_strike_count = int(pm.get("roll_chain_strike_count", 20))
                    snap = svc_live.snapshot(
                        symbols=chain_symbols,
                        chain_symbols=chain_symbols,
                        strike_count=max(4, roll_strike_count),
                    )
                    live_context = {
                        "quotes": snap.get("quotes", {}),
                        "option_chains": snap.get("option_chains", {}),
                    }
                    write_json(out_dir / "daily_roll_live_context.json", snap.get("trading_query_context", {}))
            except Exception as e:
                actions.append(f"Roll planner live-chain unavailable: {e}")
        position_decisions = build_position_decisions(pos.copy(), last_day, cfg, live_context=live_context)
        pos = pos[pos["risk"].notna() & (pos["risk"] > 0)].copy()
        if not pos.empty:
            total_risk = float(pos["risk"].sum())
            short_put_risk = float(pos.loc[pos["strategy"] == "Short Put Option", "risk"].sum())
            short_put_share = short_put_risk / total_risk if total_risk > 0 else 0.0
            pos_summary = {
                "total_open_risk": total_risk,
                "short_put_risk": short_put_risk,
                "short_put_risk_share": short_put_share,
            }

            symbol_risk_table = (
                pos.groupby("underlying", dropna=False)["risk"]
                .sum()
                .reset_index(name="risk")
                .sort_values("risk", ascending=False)
                .reset_index(drop=True)
            )
            symbol_risk_table["risk_share"] = symbol_risk_table["risk"] / max(1e-9, total_risk)

            if short_put_share > short_put_limit:
                escalate(
                    "RED",
                    f"Short-put risk share {short_put_share:.1%} > limit {short_put_limit:.1%}",
                    "Do not add short puts; trim/roll largest short-put exposure first.",
                )

            sym_breach = symbol_risk_table[symbol_risk_table["risk_share"] > symbol_limit]
            if not sym_breach.empty:
                names = ", ".join(sym_breach["underlying"].head(5).tolist())
                escalate(
                    "YELLOW",
                    f"Single-symbol risk concentration above {symbol_limit:.1%}: {names}",
                    "Reduce concentrated ticker risk before adding same-ticker trades.",
                )

            short_put_pos = pos[pos["strategy"] == "Short Put Option"].copy()
            short_put_pos = short_put_pos[short_put_pos["expiry"].notna()].copy()
            if not short_put_pos.empty and short_put_risk > 0:
                expiry_risk_table = (
                    short_put_pos.groupby(short_put_pos["expiry"].dt.date)["risk"]
                    .sum()
                    .reset_index(name="risk")
                    .sort_values("risk", ascending=False)
                    .reset_index(drop=True)
                )
                expiry_risk_table["risk_share"] = expiry_risk_table["risk"] / max(1e-9, short_put_risk)
                exp_breach = expiry_risk_table[expiry_risk_table["risk_share"] > expiry_limit]
                if not exp_breach.empty:
                    exp = str(exp_breach.iloc[0, 0])
                    share = float(exp_breach.iloc[0]["risk_share"])
                    escalate(
                        "YELLOW",
                        f"Short-put expiry concentration {share:.1%} on {exp} > {expiry_limit:.1%}",
                        "Spread short-put expiries; avoid stacking same expiration risk.",
                    )
    else:
        actions.append(
            "Close/roll analysis unavailable: provide --open-positions-csv with unrealized and risk fields."
        )

    payload = {
        "mode": "daily",
        "status": status,
        "as_of_trade_date": str(last_day),
        "day_realized_pnl": day_pnl,
        "recent_trade_count": int(len(recent)),
        "recent_net_pnl": recent_net,
        "recent_profit_factor": recent_pf,
        "trailing_loss_streak": int(cur_loss_streak),
        "thresholds": {
            "daily_loss_stop_cash": daily_loss_stop_cash,
            "loss_streak_yellow": streak_yellow,
            "loss_streak_red": streak_red,
            "rolling_pf_yellow_floor": min_pf_yellow,
            "short_put_max_share": short_put_limit,
            "single_symbol_max_share": symbol_limit,
            "single_expiry_max_share_short_put": expiry_limit,
        },
        "triggers": triggers,
        "actions": actions,
        "open_position_risk_summary": pos_summary,
        "position_decision_counts": (
            position_decisions["recommendation"].value_counts().to_dict() if not position_decisions.empty else {}
        ),
    }
    write_json(out_dir / "daily_risk_monitor.json", payload)

    lines = [
        "# Daily Risk Monitor",
        f"Status: **{status}**",
        "",
        "## Snapshot",
        f"- As-of trade date: {last_day}",
        f"- Day realized P/L: {day_pnl:,.2f}",
        f"- Recent trades checked: {len(recent)}",
        f"- Recent net P/L: {recent_net:,.2f}",
        f"- Recent PF: {recent_pf:.2f}" if np.isfinite(recent_pf) else "- Recent PF: n/a",
        f"- Trailing loss streak: {cur_loss_streak}",
        "",
        "## Triggers",
    ]
    if triggers:
        lines.extend([f"- {t}" for t in triggers])
    else:
        lines.append("- None")
    lines.append("")
    lines.append("## Actions")
    if actions:
        lines.extend([f"- {a}" for a in actions])
    else:
        lines.append("- Continue normal sizing under current limits.")
    lines.append("")
    if pos_summary:
        lines.append("## Open Risk")
        lines.append(
            f"- Total open risk: {pos_summary.get('total_open_risk', 0.0):,.2f} | "
            f"Short put share: {pos_summary.get('short_put_risk_share', 0.0):.1%}"
        )
        lines.append("")
        lines.append("### Symbol Concentration")
        lines.append(to_md_table(symbol_risk_table[["underlying", "risk", "risk_share"]], n=20))
        lines.append("")
        lines.append("### Short Put Expiry Concentration")
        if not expiry_risk_table.empty:
            lines.append(to_md_table(expiry_risk_table, n=20))
        else:
            lines.append("_none_")
        lines.append("")
    if not position_decisions.empty:
        lines.append("## Position Decisions")
        lines.append("- Recommendations are structure-level (spreads/legs grouped) and prioritize capital protection first.")
        lines.append("")
        decision_cols = [
            "underlying",
            "strategy",
            "legs",
            "expiry",
            "dte",
            "action",
            "reason",
            "roll_plan",
        ]
        for c in decision_cols:
            if c not in position_decisions.columns:
                position_decisions[c] = ""
        lines.append(
            to_md_table(
                position_decisions[decision_cols],
                n=150,
            )
        )
        lines.append("")
        close_q = position_decisions[position_decisions["recommendation"] == "CLOSE_NOW"].copy()
        roll_q = position_decisions[position_decisions["recommendation"] == "ROLL"].copy()
        wait_q = position_decisions[position_decisions["recommendation"] == "WAIT"].copy()
        data_q = position_decisions[position_decisions["recommendation"] == "DATA_NEEDED"].copy()

        lines.append("## Close Now Queue")
        if not close_q.empty:
            lines.append(
                to_md_table(
                    close_q[["underlying", "strategy", "legs", "expiry", "dte", "reason"]],
                    n=100,
                )
            )
        else:
            lines.append("_none_")
        lines.append("")

        lines.append("## Roll Queue")
        if not roll_q.empty:
            lines.append(
                to_md_table(
                    roll_q[["underlying", "strategy", "legs", "expiry", "dte", "roll_plan", "reason"]],
                    n=100,
                )
            )
        else:
            lines.append("_none_")
        lines.append("")

        lines.append("## Wait / Monitor")
        if not wait_q.empty:
            lines.append(
                to_md_table(
                    wait_q[["underlying", "strategy", "legs", "expiry", "dte", "reason"]],
                    n=150,
                )
            )
        else:
            lines.append("_none_")
        lines.append("")

        lines.append("## Data Gaps")
        if not data_q.empty:
            lines.append(
                to_md_table(
                    data_q[["underlying", "strategy", "legs", "expiry", "dte", "reason"]],
                    n=100,
                )
            )
        else:
            lines.append("_none_")
        lines.append("")
    write_text(out_dir / "daily_risk_monitor.md", "\n".join(lines))
    return payload


def aggregate_edge(df: pd.DataFrame, group_cols: List[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    g = (
        df.groupby(group_cols, dropna=False)
        .agg(
            trades=("realized_pnl", "size"),
            win_rate=("win", "mean"),
            net_pnl=("realized_pnl", "sum"),
            avg_pnl=("realized_pnl", "mean"),
            profit_factor=("realized_pnl", compute_profit_factor),
        )
        .reset_index()
        .sort_values("net_pnl", ascending=False)
        .reset_index(drop=True)
    )
    return g


def run_weekly(trades: pd.DataFrame, out_dir: Path, cfg: Dict, as_of: Optional[pd.Timestamp]) -> Dict:
    pb = cfg.get("playbook", {})
    wk = pb.get("weekly", {})
    strategy_window = int(wk.get("strategy_window_trades", 20))
    strategy_min_trades = int(wk.get("strategy_min_trades", 20))
    strategy_min_pf = float(wk.get("strategy_min_pf", 1.0))
    symbol_window = int(wk.get("symbol_window_trades", 20))
    symbol_min_trades = int(wk.get("symbol_min_trades", 5))
    symbol_pause_net = float(wk.get("symbol_pause_if_net_below", 0.0))
    combo_min_trades = int(wk.get("combo_min_trades", 3))
    combo_pause_net = float(wk.get("combo_pause_if_net_below", -1000.0))

    x = trades.sort_values("date").reset_index(drop=True)
    if as_of is not None:
        x = x[x["date"] <= as_of].copy()

    strategy_rows = []
    for strategy, g in x.groupby("strategy", dropna=False):
        t = g.tail(max(1, strategy_window)).copy()
        trades_n = int(len(t))
        pf = compute_profit_factor(t["realized_pnl"])
        net = float(t["realized_pnl"].sum())
        wr = float((t["realized_pnl"] > 0).mean())
        action = "KEEP"
        reason = ""
        if trades_n >= strategy_min_trades and np.isfinite(pf) and pf < strategy_min_pf:
            action = "PAUSE"
            reason = f"PF {pf:.2f} < {strategy_min_pf:.2f}"
        strategy_rows.append(
            {
                "strategy": strategy,
                "window_trades": trades_n,
                "window_win_rate": wr,
                "window_net_pnl": net,
                "window_profit_factor": pf,
                "action": action,
                "reason": reason,
            }
        )
    strategy_actions = pd.DataFrame(strategy_rows).sort_values(["action", "window_net_pnl"], ascending=[True, False])

    symbol_rows = []
    for symbol, g in x.groupby("symbol", dropna=False):
        t = g.tail(max(1, symbol_window)).copy()
        trades_n = int(len(t))
        net = float(t["realized_pnl"].sum())
        pf = compute_profit_factor(t["realized_pnl"])
        wr = float((t["realized_pnl"] > 0).mean())
        action = "KEEP"
        reason = ""
        if trades_n >= symbol_min_trades and net <= symbol_pause_net:
            action = "PAUSE"
            reason = f"Net {net:,.2f} <= {symbol_pause_net:,.2f}"
        symbol_rows.append(
            {
                "symbol": symbol,
                "window_trades": trades_n,
                "window_win_rate": wr,
                "window_net_pnl": net,
                "window_profit_factor": pf,
                "action": action,
                "reason": reason,
            }
        )
    symbol_actions = pd.DataFrame(symbol_rows).sort_values(["action", "window_net_pnl"], ascending=[True, False])

    combo_stats = aggregate_edge(x, ["strategy", "symbol"])
    combo_actions = combo_stats.copy()
    if not combo_actions.empty:
        combo_actions["action"] = "KEEP"
        combo_actions["reason"] = ""
        bad = (combo_actions["trades"] >= combo_min_trades) & (combo_actions["net_pnl"] <= combo_pause_net)
        combo_actions.loc[bad, "action"] = "PAUSE"
        combo_actions.loc[bad, "reason"] = combo_actions.loc[bad, "net_pnl"].map(
            lambda v: f"net_pnl {v:,.2f} <= {combo_pause_net:,.2f}"
        )
        combo_actions = combo_actions.sort_values(["action", "net_pnl"], ascending=[True, False]).reset_index(drop=True)

    status = "GREEN"
    if (strategy_actions["action"] == "PAUSE").any() or (symbol_actions["action"] == "PAUSE").any():
        status = "YELLOW"
    if not combo_actions.empty and (combo_actions["action"] == "PAUSE").any():
        status = "RED"

    payload = {
        "mode": "weekly",
        "status": status,
        "as_of_date": str(x["date"].max().date()) if not x.empty else "",
        "strategy_pauses": int((strategy_actions["action"] == "PAUSE").sum()),
        "symbol_pauses": int((symbol_actions["action"] == "PAUSE").sum()),
        "combo_pauses": int((combo_actions["action"] == "PAUSE").sum()) if not combo_actions.empty else 0,
        "thresholds": {
            "strategy_window_trades": strategy_window,
            "strategy_min_trades": strategy_min_trades,
            "strategy_min_pf": strategy_min_pf,
            "symbol_window_trades": symbol_window,
            "symbol_min_trades": symbol_min_trades,
            "symbol_pause_if_net_below": symbol_pause_net,
            "combo_min_trades": combo_min_trades,
            "combo_pause_if_net_below": combo_pause_net,
        },
    }
    write_json(out_dir / "weekly_edge_report.json", payload)
    strategy_actions.to_csv(out_dir / "weekly_strategy_actions.csv", index=False)
    symbol_actions.to_csv(out_dir / "weekly_symbol_actions.csv", index=False)
    if not combo_actions.empty:
        combo_actions.to_csv(out_dir / "weekly_combo_actions.csv", index=False)

    lines = [
        "# Weekly Edge Report",
        f"Status: **{status}**",
        "",
        "## Pause Summary",
        f"- Strategy pauses: {payload['strategy_pauses']}",
        f"- Symbol pauses: {payload['symbol_pauses']}",
        f"- Strategy-symbol combo pauses: {payload['combo_pauses']}",
        "",
        "## Strategy Actions",
        to_md_table(strategy_actions, n=50),
        "",
        "## Symbol Actions",
        to_md_table(symbol_actions, n=50),
        "",
    ]
    if not combo_actions.empty:
        lines.extend(
            [
                "## Strategy-Symbol Actions",
                to_md_table(combo_actions, n=100),
                "",
            ]
        )
    write_text(out_dir / "weekly_edge_report.md", "\n".join(lines))
    return payload


def run_monthly(trades: pd.DataFrame, out_dir: Path, cfg: Dict) -> Dict:
    pb = cfg.get("playbook", {})
    mo = pb.get("monthly", {})
    target_monthly = float(mo.get("target_monthly_profit", 20000.0))
    long_call_kill = float(mo.get("long_call_monthly_kill_switch", -1500.0))

    x = trades.sort_values("date").reset_index(drop=True)
    monthly = (
        x.groupby("month", dropna=False)
        .agg(
            trades=("realized_pnl", "size"),
            win_rate=("win", "mean"),
            net_pnl=("realized_pnl", "sum"),
            avg_pnl=("realized_pnl", "mean"),
            profit_factor=("realized_pnl", compute_profit_factor),
        )
        .reset_index()
        .sort_values("month")
        .reset_index(drop=True)
    )
    monthly["rolling_3m_net"] = monthly["net_pnl"].rolling(3, min_periods=1).mean()
    monthly["rolling_3m_pf"] = monthly["profit_factor"].rolling(3, min_periods=1).mean()

    avg_monthly = float(monthly["net_pnl"].mean()) if not monthly.empty else math.nan
    med_monthly = float(monthly["net_pnl"].median()) if not monthly.empty else math.nan
    last_3m_avg = float(monthly["net_pnl"].tail(3).mean()) if not monthly.empty else math.nan
    monthly_gap = target_monthly - avg_monthly if np.isfinite(avg_monthly) else math.nan
    multiplier = target_monthly / avg_monthly if np.isfinite(avg_monthly) and avg_monthly > 0 else math.nan

    by_strategy = aggregate_edge(x, ["strategy"])
    by_symbol = aggregate_edge(x, ["symbol"])
    by_strategy_month = (
        x.groupby(["month", "strategy"], dropna=False)["realized_pnl"]
        .sum()
        .reset_index(name="net_pnl")
        .sort_values(["month", "net_pnl"], ascending=[True, False])
        .reset_index(drop=True)
    )

    long_call_month = (
        x[x["strategy"] == "Long Call Option"]
        .groupby("month", dropna=False)["realized_pnl"]
        .sum()
        .reset_index(name="long_call_net")
        .sort_values("month")
    )
    long_call_breach = long_call_month[long_call_month["long_call_net"] <= long_call_kill].copy()

    top_losses = x.sort_values("realized_pnl").head(20).copy()

    payload = {
        "mode": "monthly",
        "as_of_date": str(x["date"].max().date()) if not x.empty else "",
        "months_analyzed": int(len(monthly)),
        "avg_monthly_net_pnl": avg_monthly,
        "median_monthly_net_pnl": med_monthly,
        "last_3m_avg_net_pnl": last_3m_avg,
        "target_monthly_profit": target_monthly,
        "gap_to_target": monthly_gap,
        "required_multiplier_to_target": multiplier,
        "long_call_kill_switch_threshold": long_call_kill,
        "long_call_breach_months": long_call_breach["month"].tolist(),
    }
    write_json(out_dir / "monthly_longitudinal_review.json", payload)
    monthly.to_csv(out_dir / "monthly_pnl_trend.csv", index=False)
    by_strategy.to_csv(out_dir / "monthly_strategy_edge.csv", index=False)
    by_symbol.to_csv(out_dir / "monthly_symbol_edge.csv", index=False)
    by_strategy_month.to_csv(out_dir / "monthly_strategy_month_matrix.csv", index=False)
    long_call_month.to_csv(out_dir / "monthly_long_call_net.csv", index=False)
    top_losses.to_csv(out_dir / "monthly_top_losses.csv", index=False)

    lines = [
        "# Monthly Longitudinal Review",
        "",
        "## Performance Baseline",
        f"- Months analyzed: {payload['months_analyzed']}",
        f"- Average monthly net P/L: {avg_monthly:,.2f}" if np.isfinite(avg_monthly) else "- Average monthly net P/L: n/a",
        f"- Median monthly net P/L: {med_monthly:,.2f}" if np.isfinite(med_monthly) else "- Median monthly net P/L: n/a",
        f"- Last 3-month avg net P/L: {last_3m_avg:,.2f}" if np.isfinite(last_3m_avg) else "- Last 3-month avg net P/L: n/a",
        f"- Target monthly net P/L: {target_monthly:,.2f}",
        f"- Gap to target: {monthly_gap:,.2f}" if np.isfinite(monthly_gap) else "- Gap to target: n/a",
        f"- Required multiplier to target: {multiplier:.2f}x" if np.isfinite(multiplier) else "- Required multiplier to target: n/a",
        "",
        "## Long Call Kill-Switch Check",
        f"- Threshold: monthly long-call net <= {long_call_kill:,.2f}",
    ]
    if not long_call_breach.empty:
        lines.append("- Breach months: " + ", ".join(long_call_breach["month"].tolist()))
    else:
        lines.append("- Breach months: none")
    lines.extend(
        [
            "",
            "## Monthly Net Trend",
            to_md_table(monthly, n=36),
            "",
            "## Strategy Edge",
            to_md_table(by_strategy, n=30),
            "",
            "## Symbol Edge",
            to_md_table(by_symbol, n=30),
            "",
            "## Largest Losses",
            to_md_table(top_losses[["date", "symbol", "strategy", "realized_pnl"]], n=20),
            "",
        ]
    )
    write_text(out_dir / "monthly_longitudinal_review.md", "\n".join(lines))
    return payload


def build_arg_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="Trade playbook pipeline: daily risk, weekly edge, monthly review.")
    ap.add_argument("mode", choices=["daily", "weekly", "monthly", "all"])
    ap.add_argument(
        "--realized-source",
        choices=["csv", "sheet", "schwab"],
        default="csv",
        help="Source of realized trade history.",
    )
    ap.add_argument(
        "--realized-csv",
        default=r"c:\uw_root\out\trade_performance_review_manual_options_full\cleaned_realized_trades.csv",
        help="CSV with realized trades (date, strategy, symbol, realized_pnl).",
    )
    ap.add_argument(
        "--sheet-csv-url",
        default="",
        help="Optional Google Sheet CSV export URL. If provided, pipeline pulls this URL and auto-builds realized CSV.",
    )
    ap.add_argument(
        "--sheet-cache-csv",
        default=DEFAULT_SHEET_CACHE_CSV,
        help="Where to cache pulled raw sheet CSV when --sheet-csv-url is used.",
    )
    ap.add_argument(
        "--sheet-realized-csv",
        default=DEFAULT_SHEET_REALIZED_CSV,
        help="Where to write standardized realized CSV when --sheet-csv-url is used.",
    )
    ap.add_argument(
        "--sheet-row-filter",
        choices=["all", "yellow"],
        default="all",
        help="When using --sheet-csv-url, optionally keep only yellow-highlighted rows (plus month header rows).",
    )
    ap.add_argument(
        "--open-positions-csv",
        default="",
        help="Optional open positions CSV for daily risk concentration checks.",
    )
    ap.add_argument(
        "--open-positions-source",
        choices=["csv", "schwab"],
        default="csv",
        help="Source for open positions in daily mode. 'schwab' pulls live positions from Schwab API.",
    )
    ap.add_argument(
        "--open-positions-cache-csv",
        default="",
        help="When --open-positions-source schwab, optional path to cache pulled positions CSV.",
    )
    ap.add_argument("--config", default=str((Path(__file__).resolve().parent / "rulebook_config.yaml")), help="YAML config path.")
    ap.add_argument("--out-dir", default=r"c:\uw_root\out\playbook", help="Output directory.")
    ap.add_argument(
        "--schwab-lookback-days",
        type=int,
        default=365,
        help="When --realized-source schwab, number of days of trade history to pull.",
    )
    ap.add_argument(
        "--schwab-realized-cache-csv",
        default="",
        help="When --realized-source schwab, optional path to cache reconstructed realized options trades.",
    )
    ap.add_argument("--lookback-trades", type=int, default=30, help="Recent-trade lookback for daily checks.")
    ap.add_argument("--start-date", default="", help="Optional filter start date YYYY-MM-DD.")
    ap.add_argument("--end-date", default="", help="Optional filter end date YYYY-MM-DD.")
    return ap


def main(argv: Optional[List[str]] = None) -> int:
    args = build_arg_parser().parse_args(argv)
    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    sheet_meta: Dict[str, float] = {}
    if args.realized_source == "sheet":
        if not args.sheet_csv_url:
            raise RuntimeError("--realized-source sheet requires --sheet-csv-url.")
        raw_cache_csv = (
            out_dir / "source_sheet_latest.csv"
            if args.sheet_cache_csv == DEFAULT_SHEET_CACHE_CSV
            else Path(args.sheet_cache_csv).resolve()
        )
        realized_csv = (
            out_dir / "cleaned_realized_trades_from_sheet.csv"
            if args.sheet_realized_csv == DEFAULT_SHEET_REALIZED_CSV
            else Path(args.sheet_realized_csv).resolve()
        )
        sheet_meta = build_realized_from_sheet_csv_url(
            args.sheet_csv_url,
            raw_cache_csv,
            realized_csv,
            row_filter=args.sheet_row_filter,
        )
    elif args.realized_source == "schwab":
        realized_csv = (
            Path(args.schwab_realized_cache_csv).resolve()
            if args.schwab_realized_cache_csv
            else (out_dir / "cleaned_realized_trades_from_schwab.csv")
        )
        sheet_meta = build_realized_from_schwab_api(realized_csv, lookback_days=int(args.schwab_lookback_days))
    else:
        realized_csv = Path(args.realized_csv).resolve()
        if not realized_csv.exists():
            raise FileNotFoundError(f"Missing realized CSV: {realized_csv}")

    cfg = load_config(Path(args.config).resolve())
    trades = load_realized_trades(realized_csv)
    if args.start_date:
        sd = dt.date.fromisoformat(args.start_date)
        trades = trades[trades["date"].dt.date >= sd].copy()
    if args.end_date:
        ed = dt.date.fromisoformat(args.end_date)
        trades = trades[trades["date"].dt.date <= ed].copy()
    if trades.empty:
        raise RuntimeError("No trades after filters.")

    daily = None
    weekly = None
    monthly = None
    open_positions_csv: Optional[Path] = Path(args.open_positions_csv).resolve() if args.open_positions_csv else None

    if args.mode in {"daily", "all"} and args.open_positions_source == "schwab":
        try:
            from uwos.schwab_auth import SchwabAuthConfig, SchwabLiveDataService
        except Exception as e:
            raise RuntimeError(f"Failed to import Schwab live service: {e}") from e

        cfg_live = SchwabAuthConfig.from_env(load_dotenv_file=True)
        svc = SchwabLiveDataService(cfg_live)
        svc.connect()
        cli = svc._client
        resp = cli.get_accounts(fields=[cli.Account.Fields.POSITIONS])
        resp.raise_for_status()
        raw = resp.json()
        accounts = raw if isinstance(raw, list) else [raw]
        acct_map_resp = cli.get_account_numbers()
        acct_map_resp.raise_for_status()
        acct_map_rows = acct_map_resp.json()
        account_number_to_hash = {
            str(a.get("accountNumber", "")).strip(): str(a.get("hashValue", "")).strip()
            for a in acct_map_rows
            if str(a.get("accountNumber", "")).strip()
        }
        open_lot_meta = build_open_option_lot_metadata_from_schwab_api(
            cli,
            account_number_to_hash,
            lookback_days=int(args.schwab_lookback_days),
        )

        rows: List[Dict[str, Any]] = []
        for a in accounts:
            sec = a.get("securitiesAccount", {}) if isinstance(a, dict) else {}
            acct_num = sec.get("accountNumber", "")
            for p in sec.get("positions", []) or []:
                inst = p.get("instrument") or {}
                rows.append(
                    {
                        "account_number": acct_num,
                        "symbol": inst.get("symbol", ""),
                        "description": inst.get("description", ""),
                        "asset_type": inst.get("assetType", ""),
                        "position_type": p.get("positionType", ""),
                        "long_quantity": p.get("longQuantity"),
                        "short_quantity": p.get("shortQuantity"),
                        "average_price": p.get("averagePrice"),
                        "market_value": p.get("marketValue"),
                        "maintenance_requirement": p.get("maintenanceRequirement"),
                        "current_day_profit_loss": p.get("currentDayProfitLoss"),
                        "current_day_profit_loss_pct": p.get("currentDayProfitLossPercentage"),
                    }
                )

        cache_csv = (
            Path(args.open_positions_cache_csv).resolve()
            if args.open_positions_cache_csv
            else (out_dir / "open_positions_from_schwab.csv")
        )
        pos_df = pd.DataFrame(
            rows,
            columns=[
                "account_number",
                "symbol",
                "description",
                "asset_type",
                "position_type",
                "long_quantity",
                "short_quantity",
                "average_price",
                "market_value",
                "maintenance_requirement",
                "current_day_profit_loss",
                "current_day_profit_loss_pct",
            ],
        )
        pos_df["position_side"] = np.where(
            pos_df["long_quantity"].map(parse_float).fillna(0.0) > 0,
            "LONG",
            np.where(pos_df["short_quantity"].map(parse_float).fillna(0.0) > 0, "SHORT", ""),
        )
        if not open_lot_meta.empty:
            pos_df = pos_df.merge(
                open_lot_meta,
                how="left",
                left_on=["account_number", "symbol", "position_side"],
                right_on=["account_number", "symbol", "side"],
            )
            pos_df = pos_df.drop(columns=["side"])
        pos_df.to_csv(cache_csv, index=False)
        open_positions_csv = cache_csv

    if args.mode in {"daily", "all"}:
        daily = run_daily(trades, out_dir, cfg, open_positions_csv, args.lookback_trades)
    if args.mode in {"weekly", "all"}:
        weekly = run_weekly(trades, out_dir, cfg, as_of=None)
    if args.mode in {"monthly", "all"}:
        monthly = run_monthly(trades, out_dir, cfg)

    if args.mode == "all":
        summary = {
            "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
            "source_realized_csv": str(realized_csv),
            "source_sheet_csv_url": args.sheet_csv_url or "",
            "daily_status": daily.get("status") if daily else "",
            "weekly_status": weekly.get("status") if weekly else "",
            "monthly_target_gap": monthly.get("gap_to_target") if monthly else math.nan,
        }
        if sheet_meta:
            summary["sheet_parser_meta"] = sheet_meta
        write_json(out_dir / "playbook_run_summary.json", summary)
        lines = [
            "# Trade Playbook Run",
            "",
            f"- Source realized CSV: `{realized_csv}`",
            (f"- Source sheet URL: `{args.sheet_csv_url}`" if args.sheet_csv_url else "- Source sheet URL: n/a"),
            f"- Daily status: **{summary['daily_status']}**",
            f"- Weekly status: **{summary['weekly_status']}**",
            (
                f"- Gap to target monthly P/L: {summary['monthly_target_gap']:,.2f}"
                if np.isfinite(summary["monthly_target_gap"])
                else "- Gap to target monthly P/L: n/a"
            ),
            "",
            "## Files",
            "- `daily_risk_monitor.md`",
            "- `weekly_edge_report.md`",
            "- `monthly_longitudinal_review.md`",
            "- `playbook_run_summary.json`",
        ]
        write_text(out_dir / "playbook_run_summary.md", "\n".join(lines))

    print(f"Mode: {args.mode}")
    print(f"Source: {realized_csv}")
    if args.realized_source == "sheet" and args.sheet_csv_url:
        print(f"Pulled sheet URL: {args.sheet_csv_url}")
        print(f"Sheet row filter: {args.sheet_row_filter}")
    if args.realized_source == "schwab":
        print(f"Realized source: Schwab API ({args.schwab_lookback_days}d lookback)")
    if args.open_positions_source == "schwab" and open_positions_csv:
        print(f"Open positions source: Schwab API (cached to {open_positions_csv})")
    print(f"Rows loaded: {len(trades)}")
    print(f"Wrote outputs to: {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())



